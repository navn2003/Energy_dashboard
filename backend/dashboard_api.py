# dashboard_api.py — Daily Forecast Analysis Dashboard router (merged from main(1).py)
# Included by main.py via: app.include_router(dashboard_router)

import csv
import io
import math
import tempfile
from datetime import date, datetime
from decimal import Decimal
from typing import Any, List, Optional, Tuple

import pymysql
from fastapi import APIRouter, Query, Request
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# ============================================================
# TARGET DB - DASHBOARD / IMPORTED FORECAST TABLES
# ============================================================

CONFIG = {
    "host": "65.1.28.178",
    "port": 3306,
    "user": "energy",
    "password": "Energy@123",
    "database": "energy_monitor",
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}


# ============================================================
# SOURCE DB - SCADA + LIVE FORECAST STATUS
# ============================================================

SOURCE_DB = {
    "host": "13.205.184.74",
    "port": 3306,
    "user": "normal_access",
    "password": "energyX@123#",
    "database": "engxai_fs",
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}


router = APIRouter()


FORECAST_TABLES = {
    "IPSL": "dss_forecast_ipsl",
    "QueNext": "dss_forecast_quenext",
}


# ============================================================
# CONNECTIONS
# ============================================================

def get_conn():
    return pymysql.connect(**CONFIG)


def get_source_conn():
    return pymysql.connect(**SOURCE_DB)


# ============================================================
# COMMON HELPERS
# ============================================================

def safe_float(value, default=0.0):
    if value is None:
        return default

    if isinstance(value, Decimal):
        return float(value)

    try:
        value = float(value)
        if math.isnan(value):
            return default
        return value
    except Exception:
        return default


def safe_int(value, default=0):
    if value is None:
        return default

    try:
        return int(value)
    except Exception:
        return default


def table_exists(cur, table_name: str) -> bool:
    try:
        cur.execute("SHOW TABLES LIKE %s", [table_name])
        return cur.fetchone() is not None
    except Exception:
        return False


def source_table_exists(source_cur, table_name: str) -> bool:
    try:
        source_cur.execute("SHOW TABLES LIKE %s", [table_name])
        return source_cur.fetchone() is not None
    except Exception:
        return False


def get_table_columns(cur, table_name: str) -> List[str]:
    try:
        cur.execute(f"SHOW COLUMNS FROM `{table_name}`")
        return [r["Field"] for r in cur.fetchall()]
    except Exception:
        return []


def pick_col(cur, table_name: str, options: List[str]) -> Optional[str]:
    cols = get_table_columns(cur, table_name)
    lower_map = {c.lower(): c for c in cols}

    for opt in options:
        if opt.lower() in lower_map:
            return lower_map[opt.lower()]

    return None


def get_company_list(company: str) -> List[str]:
    if company == "IPSL":
        return ["IPSL"]

    if company == "QueNext":
        return ["QueNext"]

    return ["IPSL", "QueNext"]


def revision_filter(alias: str, revision: str) -> Tuple[str, List[Any]]:
    if revision == "ALL":
        return "", []

    return f" AND {alias}.revision_number = %s ", [revision]


def nrmse_status(value):
    v = safe_float(value)

    if v <= 5:
        return "GOOD"

    if v <= 9:
        return "AVERAGE"

    return "POOR"


def pay_percent_from_nrmse(revision: str, nrmse: float):
    rev = str(revision).upper().strip()

    if rev in ["R-16", "R16", "R20"]:
        if 1 <= nrmse <= 3:
            return 25.0
        if 3 < nrmse <= 5:
            return 12.5
        return 0.0

    if rev in ["RD2", "R02", "DA2"]:
        if 0 <= nrmse <= 5:
            return 25.0
        if 5 < nrmse <= 10:
            return 12.5
        return 0.0

    return 0.0


def get_master_columns(cur):
    """
    Detect plant master table in TARGET DB.
    Used for SOLAR/WIND forecast alerts.
    """

    if table_exists(cur, "plant_master"):
        table = "plant_master"
    elif table_exists(cur, "DSS_MASTER"):
        table = "DSS_MASTER"
    else:
        return None

    dss_col = pick_col(cur, table, ["dss_id", "DSS_ID"])
    energy_col = pick_col(
        cur,
        table,
        ["energy_type", "plant_type", "ENERGY_TYPE", "PLANT_TYPE"]
    )

    if not dss_col or not energy_col:
        return None

    return {
        "table": table,
        "dss_col": dss_col,
        "energy_col": energy_col,
    }


# ============================================================
# ROUTES
# ============================================================

@router.get("/api/dashboard")
def dashboard_api(
    report_date: str = Query(default=None),
    company: str = Query(default="ALL"),
    revision: str = Query(default="ALL"),
):
    if not report_date:
        report_date = str(date.today())

    try:
        conn = get_conn()

        with conn.cursor() as cur:
            summary = get_forecast_summary(cur, report_date, company, revision)
            forecast_curve = get_forecast_curve(cur, report_date, company, revision)
            actual_curve = get_actual_curve(report_date)

            daily_nrmse = get_daily_nrmse(cur, report_date, company, revision)
            monthly_nrmse = get_monthly_nrmse(cur, report_date, company, revision)

            heatmap = get_error_heatmap(forecast_curve, actual_curve)
            revision_summary = get_revision_summary(forecast_curve)
            bias = get_bias_analysis(forecast_curve, actual_curve)

            plant_errors = get_plant_forecast_ranking(cur, report_date, company, revision)
            zero_plants = get_zero_forecast_plants(cur, report_date, company, revision)
            missing_blocks = get_missing_blocks(cur, report_date, company, revision)

            data_availability = get_data_availability(cur, report_date)
            scada_status = get_scada_data_status(report_date)

            payment_impact = get_payment_impact(daily_nrmse)

            solar_zero_blockwise = get_solar_zero_forecast_blockwise(
                cur, report_date, company, revision
            )
            wind_zero_blockwise = get_wind_zero_forecast_blockwise(
                cur, report_date, company, revision
            )

            repeated_forecast = get_repeated_forecast_count(
                cur, report_date, company, revision
            )

            solar_scada_night = get_solar_scada_night_alert(report_date)
            model_status = get_forecast_model_status(report_date, revision)

            alerts = build_alerts(
                daily_nrmse=daily_nrmse,
                zero_plants=zero_plants,
                missing_blocks=missing_blocks,
                data_availability=data_availability,
                solar_zero_blockwise=solar_zero_blockwise,
                wind_zero_blockwise=wind_zero_blockwise,
                repeated_forecast=repeated_forecast,
                solar_scada_night=solar_scada_night,
                model_status=model_status,
                scada_status=scada_status,
            )

            kpis = build_kpis(
                summary=summary,
                daily_nrmse=daily_nrmse,
                zero_plants=zero_plants,
                missing_blocks=missing_blocks,
                bias=bias,
            )

        conn.close()

        return JSONResponse(
            content=jsonable_encoder(
                {
                    "error": False,
                    "report_date": report_date,
                    "company": company,
                    "revision": revision,
                    "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "kpis": kpis,
                    "summary": summary,
                    "forecast_curve": forecast_curve,
                    "actual_curve": actual_curve,
                    "daily_nrmse": daily_nrmse,
                    "monthly_nrmse": monthly_nrmse,
                    "heatmap": heatmap,
                    "revision_summary": revision_summary,
                    "bias": bias,
                    "plant_errors": plant_errors,
                    "zero_plants": zero_plants,
                    "missing_blocks": missing_blocks,
                    "data_availability": data_availability,
                    "scada_status": scada_status,
                    "payment_impact": payment_impact,
                    "model_status": model_status,
                    "solar_zero_blockwise": solar_zero_blockwise,
                    "wind_zero_blockwise": wind_zero_blockwise,
                    "alerts": alerts,
                }
            )
        )

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "error": True,
                "message": "Dashboard API failed",
                "detail": str(e),
            },
        )


# ============================================================
# FORECAST DATA FROM TARGET DB
# ============================================================

def get_forecast_summary(cur, report_date, company, revision):
    rows = []

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                %s AS company,
                f.revision_number,
                f.forecast_date,
                COUNT(*) AS total_rows,
                COUNT(DISTINCT f.dss_id) AS dss_count,
                COUNT(DISTINCT f.block_no) AS block_count,
                ROUND(SUM(f.forecast_mw), 3) AS total_forecast_mw,
                ROUND(AVG(f.forecast_mw), 3) AS avg_forecast_mw,
                ROUND(MAX(f.forecast_mw), 3) AS max_forecast_mw,
                SUM(CASE WHEN IFNULL(f.forecast_mw, 0) = 0 THEN 1 ELSE 0 END) AS zero_rows,
                MAX(f.updated_at) AS latest_updated_at
            FROM `{table}` f
            WHERE f.forecast_date = %s
            {rev_sql}
            GROUP BY f.revision_number, f.forecast_date
            ORDER BY f.revision_number
        """

        cur.execute(sql, [comp, report_date] + rev_params)
        rows.extend(cur.fetchall())

    return rows


def get_forecast_curve(cur, report_date, company, revision):
    """
    Forecast curve:
    Adds all plant forecast_mw for each block.
    """

    rows = []

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                %s AS company,
                f.revision_number,
                f.block_no,
                f.block_time,
                ROUND(SUM(f.forecast_mw), 3) AS forecast_mw
            FROM `{table}` f
            WHERE f.forecast_date = %s
            {rev_sql}
            GROUP BY f.revision_number, f.block_no, f.block_time
            ORDER BY f.revision_number, f.block_no
        """

        cur.execute(sql, [comp, report_date] + rev_params)
        rows.extend(cur.fetchall())

    return rows


# ============================================================
# SCADA ACTUAL FROM SOURCE DB - 13.205.184.74
# ============================================================

def get_actual_curve(report_date):
    """
    SCADA actual curve from SOURCE DB:
        Host   : 13.205.184.74
        DB     : engxai_fs
        Table  : DSS_ACTUAL
        Column : SCADA_POWER_MW

    Adds all DSS plant SCADA for each 15-minute timestamp.
    """

    source_conn = None

    try:
        source_conn = get_source_conn()

        with source_conn.cursor() as source_cur:
            sql = """
                SELECT
                    FLOOR((HOUR(`TIMESTAMP`) * 60 + MINUTE(`TIMESTAMP`)) / 15) + 1 AS block_no,
                    TIME(`TIMESTAMP`) AS block_time,
                    ROUND(SUM(SCADA_POWER_MW), 3) AS actual_mw
                FROM DSS_ACTUAL
                WHERE DATE(`TIMESTAMP`) = %s
                  AND SCADA_POWER_MW IS NOT NULL
                GROUP BY block_no, TIME(`TIMESTAMP`)
                ORDER BY block_no
            """

            source_cur.execute(sql, [report_date])
            rows = source_cur.fetchall()

        return rows

    except Exception as e:
        print("SCADA source DB error:", e)
        return []

    finally:
        if source_conn:
            source_conn.close()


def get_scada_data_status(report_date):
    """
    SCADA availability alert from SOURCE DB:
        Host   : 13.205.184.74
        DB     : engxai_fs
        Table  : DSS_ACTUAL
    """

    result = {
        "status": "NOT_AVAILABLE",
        "rows": 0,
        "dss_count": 0,
        "from_time": None,
        "to_time": None,
        "total_scada": 0,
    }

    source_conn = None

    try:
        source_conn = get_source_conn()

        with source_conn.cursor() as source_cur:
            sql = """
                SELECT
                    COUNT(*) AS rows_count,
                    COUNT(DISTINCT DSS_ID) AS dss_count,
                    MIN(`TIMESTAMP`) AS from_time,
                    MAX(`TIMESTAMP`) AS to_time,
                    ROUND(SUM(SCADA_POWER_MW), 3) AS total_scada
                FROM DSS_ACTUAL
                WHERE DATE(`TIMESTAMP`) = %s
                  AND SCADA_POWER_MW IS NOT NULL
            """

            source_cur.execute(sql, [report_date])
            row = source_cur.fetchone() or {}

        rows_count = safe_int(row.get("rows_count"))

        result["rows"] = rows_count
        result["dss_count"] = safe_int(row.get("dss_count"))
        result["from_time"] = str(row.get("from_time")) if row.get("from_time") else None
        result["to_time"] = str(row.get("to_time")) if row.get("to_time") else None
        result["total_scada"] = safe_float(row.get("total_scada"))

        if rows_count > 0:
            result["status"] = "AVAILABLE"

        return result

    except Exception as e:
        result["status"] = "ERROR"
        result["error"] = str(e)
        return result

    finally:
        if source_conn:
            source_conn.close()


def get_data_availability(cur, report_date):
    result = {
        "forecast_data": 0,
        "scada_actuals": 0,
    }

    forecast_count = 0

    for table in FORECAST_TABLES.values():
        if table_exists(cur, table):
            cur.execute(
                f"""
                SELECT COUNT(*) AS cnt
                FROM `{table}`
                WHERE forecast_date = %s
                """,
                [report_date],
            )

            row = cur.fetchone()
            forecast_count += safe_int(row.get("cnt") if row else 0)

    result["forecast_data"] = 100 if forecast_count > 0 else 0

    source_conn = None

    try:
        source_conn = get_source_conn()

        with source_conn.cursor() as source_cur:
            source_cur.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM DSS_ACTUAL
                WHERE DATE(`TIMESTAMP`) = %s
                  AND SCADA_POWER_MW IS NOT NULL
                """,
                [report_date],
            )

            row = source_cur.fetchone()
            scada_count = safe_int(row.get("cnt") if row else 0)

        result["scada_actuals"] = 100 if scada_count > 0 else 0

    except Exception as e:
        print("SCADA availability source DB error:", e)
        result["scada_actuals"] = 0

    finally:
        if source_conn:
            source_conn.close()

    return result


# ============================================================
# NRMSE / SUMMARY
# ============================================================

def get_daily_nrmse(cur, report_date, company, revision):
    if not table_exists(cur, "daily_nrmse_values"):
        return []

    sql = """
        SELECT
            date,
            company,
            energy_type,
            revision_num,
            ROUND(nrmse_values, 4) AS nrmse_values,
            updated_timestamp
        FROM daily_nrmse_values
        WHERE date = %s
    """

    params = [report_date]

    if company != "ALL":
        sql += " AND LOWER(company) = LOWER(%s)"
        params.append(company)

    if revision != "ALL":
        sql += " AND revision_num = %s"
        params.append(revision)

    sql += " ORDER BY company, energy_type, revision_num"

    cur.execute(sql, params)
    rows = cur.fetchall()

    for r in rows:
        r["status"] = nrmse_status(r.get("nrmse_values"))

    return rows


def get_monthly_nrmse(cur, report_date, company, revision):
    if not table_exists(cur, "monthly_nrmse_values"):
        return []

    month_value = report_date[:7] + "-01"

    sql = """
        SELECT
            month,
            month_name,
            company,
            energy_type,
            revision_num,
            ROUND(nrmse_values, 4) AS nrmse_values,
            updated_timestamp
        FROM monthly_nrmse_values
        WHERE month = %s
    """

    params = [month_value]

    if company != "ALL":
        sql += " AND LOWER(company) = LOWER(%s)"
        params.append(company)

    if revision != "ALL":
        sql += " AND revision_num = %s"
        params.append(revision)

    sql += " ORDER BY company, energy_type, revision_num"

    cur.execute(sql, params)
    rows = cur.fetchall()

    for r in rows:
        r["status"] = nrmse_status(r.get("nrmse_values"))

    return rows


def get_error_heatmap(forecast_curve, actual_curve):
    actual_map = {
        safe_int(r.get("block_no")): safe_float(r.get("actual_mw"))
        for r in actual_curve
    }

    result = []

    for r in forecast_curve:
        block_no = safe_int(r.get("block_no"))
        forecast_mw = safe_float(r.get("forecast_mw"))
        actual_mw = actual_map.get(block_no, 0.0)

        if actual_mw > 0:
            error_pct = ((forecast_mw - actual_mw) / actual_mw) * 100
        else:
            error_pct = 0.0

        result.append(
            {
                "company": r.get("company"),
                "revision_number": r.get("revision_number"),
                "block_no": block_no,
                "block_time": str(r.get("block_time")),
                "error_pct": round(error_pct, 2),
            }
        )

    return result


def get_revision_summary(forecast_curve):
    grouped = {}

    for r in forecast_curve:
        key = f"{r.get('company')}|{r.get('revision_number')}"

        if key not in grouped:
            grouped[key] = {
                "company": r.get("company"),
                "revision_number": r.get("revision_number"),
                "total_mw": 0.0,
                "count": 0,
                "peak_mw": 0.0,
                "peak_block": None,
            }

        forecast_mw = safe_float(r.get("forecast_mw"))
        grouped[key]["total_mw"] += forecast_mw
        grouped[key]["count"] += 1

        if forecast_mw > grouped[key]["peak_mw"]:
            grouped[key]["peak_mw"] = forecast_mw
            grouped[key]["peak_block"] = r.get("block_no")

    rows = []

    for g in grouped.values():
        avg_mw = g["total_mw"] / g["count"] if g["count"] else 0

        rows.append(
            {
                "company": g["company"],
                "revision_number": g["revision_number"],
                "total_mw": round(g["total_mw"], 3),
                "avg_mw": round(avg_mw, 3),
                "peak_mw": round(g["peak_mw"], 3),
                "peak_block": g["peak_block"],
            }
        )

    return rows


def get_bias_analysis(forecast_curve, actual_curve):
    actual_map = {
        safe_int(r.get("block_no")): safe_float(r.get("actual_mw"))
        for r in actual_curve
    }

    over_forecast = 0.0
    under_forecast = 0.0
    net_bias = 0.0

    for r in forecast_curve:
        block_no = safe_int(r.get("block_no"))
        forecast_mw = safe_float(r.get("forecast_mw"))
        actual_mw = actual_map.get(block_no, 0.0)

        diff = forecast_mw - actual_mw
        net_bias += diff

        if diff >= 0:
            over_forecast += diff
        else:
            under_forecast += abs(diff)

    total_abs = over_forecast + under_forecast

    return {
        "over_forecast_mw": round(over_forecast, 3),
        "under_forecast_mw": round(under_forecast, 3),
        "net_bias_mw": round(net_bias, 3),
        "over_pct": round((over_forecast / total_abs) * 100, 2) if total_abs else 0,
        "under_pct": round((under_forecast / total_abs) * 100, 2) if total_abs else 0,
    }


# ============================================================
# TABLE DATA
# ============================================================

def get_plant_forecast_ranking(cur, report_date, company, revision):
    rows = []

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                %s AS company,
                f.dss_id,
                f.revision_number,
                ROUND(SUM(f.forecast_mw), 3) AS forecast_mw,
                'FORECAST_ONLY' AS status
            FROM `{table}` f
            WHERE f.forecast_date = %s
            {rev_sql}
            GROUP BY f.dss_id, f.revision_number
            ORDER BY forecast_mw DESC
            LIMIT 15
        """

        cur.execute(sql, [comp, report_date] + rev_params)
        rows.extend(cur.fetchall())

    return rows[:15]


def get_zero_forecast_plants(cur, report_date, company, revision):
    """
    FULL ZERO plants only.

    A plant/revision is counted only when:
        total available blocks = 96
        zero blocks = 96
    """

    rows = []

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                %s AS company,
                f.dss_id,
                f.revision_number,
                COUNT(DISTINCT f.block_no) AS total_blocks,
                SUM(CASE WHEN IFNULL(f.forecast_mw, 0) = 0 THEN 1 ELSE 0 END) AS zero_blocks,
                ROUND(SUM(f.forecast_mw), 3) AS total_forecast_mw,
                'FULL ZERO' AS status
            FROM `{table}` f
            WHERE f.forecast_date = %s
            {rev_sql}
            GROUP BY f.dss_id, f.revision_number
            HAVING total_blocks = 96
               AND zero_blocks = 96
            ORDER BY f.dss_id, f.revision_number
            LIMIT 500
        """

        cur.execute(sql, [comp, report_date] + rev_params)
        rows.extend(cur.fetchall())

    return rows


def get_missing_blocks(cur, report_date, company, revision):
    rows = []

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                %s AS company,
                f.dss_id,
                f.revision_number,
                COUNT(DISTINCT f.block_no) AS available_blocks,
                96 - COUNT(DISTINCT f.block_no) AS missing_blocks,
                ROUND(((96 - COUNT(DISTINCT f.block_no)) / 96) * 100, 2) AS missing_pct
            FROM `{table}` f
            WHERE f.forecast_date = %s
            {rev_sql}
            GROUP BY f.dss_id, f.revision_number
            HAVING missing_blocks > 0
            ORDER BY missing_blocks DESC
            LIMIT 100
        """

        cur.execute(sql, [comp, report_date] + rev_params)
        rows.extend(cur.fetchall())

    return rows


def get_payment_impact(daily_nrmse):
    rows = []

    for r in daily_nrmse:
        nrmse = safe_float(r.get("nrmse_values"))
        rev = r.get("revision_num")
        pay = pay_percent_from_nrmse(rev, nrmse)

        rows.append(
            {
                "company": r.get("company"),
                "energy_type": r.get("energy_type"),
                "revision_num": rev,
                "nrmse_values": round(nrmse, 4),
                "eligible_pay": pay,
                "lost_pay": round(25.0 - pay, 2),
                "reason": nrmse_status(nrmse),
            }
        )

    return rows


# ============================================================
# ALERT FUNCTIONS
# ============================================================

def get_forecast_model_status(report_date, revision):
    """
    Forecast model status from SOURCE DB:
        Host   : 13.205.184.74
        DB     : engxai_fs
        Table  : DSS_FORECAST
        Field  : UPDATED_TIMESTAMP

    Rule:
        Model should update every 90 minutes.
    """

    source_conn = None

    try:
        source_conn = get_source_conn()

        with source_conn.cursor() as source_cur:
            if not source_table_exists(source_cur, "DSS_FORECAST"):
                return {
                    "status": "ERROR",
                    "message": "Source DSS_FORECAST table not found in 13.205.184.74 / engxai_fs.",
                    "total_rows": 0,
                    "latest_update": None,
                }

            sql = """
                SELECT
                    COUNT(*) AS total_rows,
                    MAX(UPDATED_TIMESTAMP) AS latest_update,
                    MAX(TIMESTAMP) AS latest_forecast_timestamp,
                    COUNT(DISTINCT DSS_ID) AS dss_count,
                    COUNT(DISTINCT REVISION_NUMBER) AS revision_count
                FROM DSS_FORECAST
                WHERE DATE(TIMESTAMP) = %s
            """

            params = [report_date]

            if revision != "ALL":
                sql += " AND REVISION_NUMBER = %s"
                params.append(revision)

            source_cur.execute(sql, params)
            row = source_cur.fetchone() or {}

        total_rows = safe_int(row.get("total_rows"))
        dss_count = safe_int(row.get("dss_count"))
        revision_count = safe_int(row.get("revision_count"))
        latest_update = row.get("latest_update")
        latest_forecast_timestamp = row.get("latest_forecast_timestamp")

        if total_rows == 0:
            return {
                "status": "NOT_RUNNING",
                "message": "Forecast model is not running. No rows found in source DSS_FORECAST for selected date.",
                "total_rows": 0,
                "dss_count": 0,
                "revision_count": 0,
                "latest_update": None,
                "latest_forecast_timestamp": None,
            }

        if latest_update is None:
            return {
                "status": "STALE",
                "message": f"Forecast rows found, but UPDATED_TIMESTAMP is missing. Rows: {total_rows}.",
                "total_rows": total_rows,
                "dss_count": dss_count,
                "revision_count": revision_count,
                "latest_update": None,
                "latest_forecast_timestamp": str(latest_forecast_timestamp),
            }

        if isinstance(latest_update, str):
            latest_dt = datetime.strptime(str(latest_update), "%Y-%m-%d %H:%M:%S")
        else:
            latest_dt = latest_update

        diff_minutes = (datetime.now() - latest_dt).total_seconds() / 60

        if diff_minutes <= 90:
            return {
                "status": "RUNNING",
                "message": (
                    f"Forecast model is running. Latest update: {latest_dt}. "
                    f"Updated {diff_minutes:.0f} minutes ago. "
                    f"Rows: {total_rows}, DSS: {dss_count}, revisions: {revision_count}."
                ),
                "total_rows": total_rows,
                "dss_count": dss_count,
                "revision_count": revision_count,
                "latest_update": str(latest_dt),
                "latest_forecast_timestamp": str(latest_forecast_timestamp),
            }

        return {
            "status": "STALE",
            "message": (
                f"Forecast model may not be running. Latest update: {latest_dt}. "
                f"Last update is {diff_minutes:.0f} minutes old, more than 90 minutes."
            ),
            "total_rows": total_rows,
            "dss_count": dss_count,
            "revision_count": revision_count,
            "latest_update": str(latest_dt),
            "latest_forecast_timestamp": str(latest_forecast_timestamp),
        }

    except Exception as e:
        return {
            "status": "ERROR",
            "message": f"Forecast model status check failed from source DSS_FORECAST: {e}",
            "total_rows": 0,
            "latest_update": None,
        }

    finally:
        if source_conn:
            source_conn.close()


def get_solar_zero_forecast_blockwise(cur, report_date, company, revision):
    """
    Blockwise solar zero forecast count.
    Checks SOLAR plants between 05:30 and 19:30.
    For each block, counts unique DSS_ID where forecast_mw = 0.
    """

    master = get_master_columns(cur)

    if not master:
        return []

    block_map = {}

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                f.block_no,
                f.block_time,
                f.dss_id
            FROM `{table}` f
            JOIN `{master["table"]}` m
                ON m.`{master["dss_col"]}` = f.dss_id
            WHERE f.forecast_date = %s
              AND UPPER(TRIM(m.`{master["energy_col"]}`)) = 'SOLAR'
              AND f.block_time BETWEEN '05:30:00' AND '19:30:00'
              AND IFNULL(f.forecast_mw, 0) = 0
              {rev_sql}
            ORDER BY f.block_no
        """

        cur.execute(sql, [report_date] + rev_params)
        rows = cur.fetchall()

        for r in rows:
            block_no = safe_int(r.get("block_no"))
            block_time = str(r.get("block_time"))
            dss_id = str(r.get("dss_id")).strip()

            key = (block_no, block_time)

            if key not in block_map:
                block_map[key] = set()

            block_map[key].add(dss_id)

    result = []

    for (block_no, block_time), dss_set in sorted(block_map.items()):
        result.append({
            "block_no": block_no,
            "block_time": block_time,
            "zero_plants": len(dss_set),
        })

    return result


def get_wind_zero_forecast_blockwise(cur, report_date, company, revision):
    """
    Blockwise wind zero forecast count.
    Checks WIND plants for all 96 blocks.
    For each block, counts unique DSS_ID where forecast_mw = 0.
    """

    master = get_master_columns(cur)

    if not master:
        return []

    block_map = {}

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                f.block_no,
                f.block_time,
                f.dss_id
            FROM `{table}` f
            JOIN `{master["table"]}` m
                ON m.`{master["dss_col"]}` = f.dss_id
            WHERE f.forecast_date = %s
              AND UPPER(TRIM(m.`{master["energy_col"]}`)) = 'WIND'
              AND IFNULL(f.forecast_mw, 0) = 0
              {rev_sql}
            ORDER BY f.block_no
        """

        cur.execute(sql, [report_date] + rev_params)
        rows = cur.fetchall()

        for r in rows:
            block_no = safe_int(r.get("block_no"))
            block_time = str(r.get("block_time"))
            dss_id = str(r.get("dss_id")).strip()

            key = (block_no, block_time)

            if key not in block_map:
                block_map[key] = set()

            block_map[key].add(dss_id)

    result = []

    for (block_no, block_time), dss_set in sorted(block_map.items()):
        result.append({
            "block_no": block_no,
            "block_time": block_time,
            "zero_plants": len(dss_set),
        })

    return result


def make_blockwise_alert_chunks(title, rows):
    """
    Creates ONE single alert line for blockwise zero forecast.
    This prevents same title appearing 4-5 times.
    """

    if not rows:
        return [f"{title}: No zero forecast blocks found."]

    parts = []

    for r in rows:
        block_time = str(r.get("block_time", ""))[:5]
        zero_plants = safe_int(r.get("zero_plants"))
        parts.append(f"{block_time}={zero_plants}")

    return [f"{title}: " + ", ".join(parts)]


def get_repeated_forecast_count(cur, report_date, company, revision):
    repeated_dss = set()
    repeated_pairs = 0

    for comp in get_company_list(company):
        table = FORECAST_TABLES[comp]

        if not table_exists(cur, table):
            continue

        rev_sql, rev_params = revision_filter("f", revision)

        sql = f"""
            SELECT
                f.dss_id,
                f.revision_number,
                f.block_no,
                ROUND(f.forecast_mw, 4) AS forecast_mw
            FROM `{table}` f
            WHERE f.forecast_date = %s
              {rev_sql}
            ORDER BY f.dss_id, f.revision_number, f.block_no
        """

        cur.execute(sql, [report_date] + rev_params)
        rows = cur.fetchall()

        last_key = None
        last_block = None
        last_value = None

        for r in rows:
            dss_id = str(r["dss_id"]).strip()
            key = f"{comp}|{dss_id}|{r['revision_number']}"

            block_no = safe_int(r.get("block_no"))
            value = round(safe_float(r.get("forecast_mw")), 4)

            if (
                last_key == key
                and last_block is not None
                and block_no == last_block + 1
                and value == last_value
            ):
                repeated_dss.add(dss_id)
                repeated_pairs += 1

            last_key = key
            last_block = block_no
            last_value = value

    return {
        "plants": len(repeated_dss),
        "pairs": repeated_pairs,
    }


def get_solar_scada_night_alert(report_date):
    """
    Solar night SCADA from SOURCE DB:
        Host   : 13.205.184.74
        DB     : engxai_fs
        Table  : DSS_ACTUAL

    If DSS_MASTER exists in source DB, it filters ENERGY_TYPE = SOLAR.
    Otherwise fallback uses all SCADA.
    """

    result = {
        "night_blocks": 0,
        "max_night_scada": 0,
        "total_night_scada": 0,
    }

    source_conn = None

    try:
        source_conn = get_source_conn()

        with source_conn.cursor() as source_cur:
            has_master = source_table_exists(source_cur, "DSS_MASTER")

            if has_master:
                sql = """
                    SELECT
                        COUNT(*) AS night_blocks,
                        ROUND(MAX(total_scada), 3) AS max_night_scada,
                        ROUND(SUM(total_scada), 3) AS total_night_scada
                    FROM (
                        SELECT
                            a.`TIMESTAMP`,
                            SUM(a.SCADA_POWER_MW) AS total_scada
                        FROM DSS_ACTUAL a
                        JOIN DSS_MASTER m
                            ON m.DSS_ID = a.DSS_ID
                        WHERE DATE(a.`TIMESTAMP`) = %s
                          AND UPPER(TRIM(m.ENERGY_TYPE)) = 'SOLAR'
                          AND a.SCADA_POWER_MW IS NOT NULL
                          AND a.SCADA_POWER_MW > 0
                          AND (
                                TIME(a.`TIMESTAMP`) < '05:30:00'
                                OR TIME(a.`TIMESTAMP`) > '19:00:00'
                              )
                        GROUP BY a.`TIMESTAMP`
                    ) x
                """

            else:
                sql = """
                    SELECT
                        COUNT(*) AS night_blocks,
                        ROUND(MAX(total_scada), 3) AS max_night_scada,
                        ROUND(SUM(total_scada), 3) AS total_night_scada
                    FROM (
                        SELECT
                            `TIMESTAMP`,
                            SUM(SCADA_POWER_MW) AS total_scada
                        FROM DSS_ACTUAL
                        WHERE DATE(`TIMESTAMP`) = %s
                          AND SCADA_POWER_MW IS NOT NULL
                          AND SCADA_POWER_MW > 0
                          AND (
                                TIME(`TIMESTAMP`) < '05:30:00'
                                OR TIME(`TIMESTAMP`) > '19:00:00'
                              )
                        GROUP BY `TIMESTAMP`
                    ) x
                """

            source_cur.execute(sql, [report_date])
            row = source_cur.fetchone() or {}

        result["night_blocks"] = safe_int(row.get("night_blocks"))
        result["max_night_scada"] = safe_float(row.get("max_night_scada"))
        result["total_night_scada"] = safe_float(row.get("total_night_scada"))

        return result

    except Exception as e:
        print("Solar night SCADA source DB error:", e)
        return result

    finally:
        if source_conn:
            source_conn.close()


def build_alerts(
    daily_nrmse,
    zero_plants,
    missing_blocks,
    data_availability,
    solar_zero_blockwise=None,
    wind_zero_blockwise=None,
    repeated_forecast=None,
    solar_scada_night=None,
    model_status=None,
    scada_status=None,
):
    alerts = []

    solar_zero_blockwise = solar_zero_blockwise or []
    wind_zero_blockwise = wind_zero_blockwise or []
    repeated_forecast = repeated_forecast or {}
    solar_scada_night = solar_scada_night or {}
    model_status = model_status or {}
    scada_status = scada_status or {}

    model_status_value = model_status.get("status", "UNKNOWN")
    model_message = model_status.get("message", "Forecast model status unknown.")

    if model_status_value in ["RUNNING", "DATA_AVAILABLE"]:
        alerts.append(f"Forecast status: {model_message}")
    else:
        alerts.append(f"Forecast status warning: {model_message}")

    if scada_status.get("status") == "AVAILABLE":
        alerts.append(
            f"SCADA data is available from source DSS_ACTUAL. "
            f"Rows: {safe_int(scada_status.get('rows'))}, "
            f"DSS count: {safe_int(scada_status.get('dss_count'))}, "
            f"From: {scada_status.get('from_time')}, "
            f"To: {scada_status.get('to_time')}, "
            f"Total SCADA: {safe_float(scada_status.get('total_scada')):.2f} MW."
        )
    elif scada_status.get("status") == "ERROR":
        alerts.append(
            f"SCADA data status check failed from source DSS_ACTUAL: {scada_status.get('error')}"
        )
    else:
        alerts.append(
            "SCADA data is not available in source DSS_ACTUAL for selected date."
        )

    alerts.extend(
        make_blockwise_alert_chunks(
            "Solar zero forecast blockwise 05:30 to 19:30",
            solar_zero_blockwise,
        )
    )

    alerts.extend(
        make_blockwise_alert_chunks(
            "Wind zero forecast blockwise all 96 blocks",
            wind_zero_blockwise,
        )
    )

    repeated_count = safe_int(repeated_forecast.get("plants"))
    repeated_pairs = safe_int(repeated_forecast.get("pairs"))

    alerts.append(
        f"{repeated_count} plants have forecast value repeated for 2 continuous blocks. "
        f"Repeated pairs: {repeated_pairs}."
    )

    solar_night_blocks = safe_int(solar_scada_night.get("night_blocks"))
    solar_night_max = safe_float(solar_scada_night.get("max_night_scada"))
    solar_night_total = safe_float(solar_scada_night.get("total_night_scada"))

    if solar_night_blocks > 0:
        alerts.append(
            f"Solar SCADA is present at night in {solar_night_blocks} blocks. "
            f"Max night SCADA: {solar_night_max:.2f} MW, "
            f"Total night SCADA: {solar_night_total:.2f} MW."
        )
    else:
        alerts.append("No solar SCADA found at night.")

    for r in daily_nrmse:
        nrmse = safe_float(r.get("nrmse_values"))

        if nrmse > 10:
            alerts.append(
                f"{r.get('company')} {r.get('energy_type')} {r.get('revision_num')} "
                f"NRMSE is {nrmse:.2f}%, above 10% threshold."
            )

    if missing_blocks:
        alerts.append(f"{len(missing_blocks)} plants have missing forecast blocks.")

    if data_availability.get("forecast_data", 0) == 0:
        alerts.append("Forecast data not available for selected date.")

    return alerts


# ============================================================
# KPI
# ============================================================

def build_kpis(summary, daily_nrmse, zero_plants, missing_blocks, bias):
    total_records = sum(safe_int(r.get("total_rows")) for r in summary)
    total_dss = max([safe_int(r.get("dss_count")) for r in summary], default=0)
    block_count = max([safe_int(r.get("block_count")) for r in summary], default=0)

    best = None
    worst = None

    if daily_nrmse:
        best = min(daily_nrmse, key=lambda x: safe_float(x.get("nrmse_values")))
        worst = max(daily_nrmse, key=lambda x: safe_float(x.get("nrmse_values")))

    full_zero_unique_dss = len(
        set(str(r.get("dss_id")).strip() for r in zero_plants)
    )

    return {
        "total_records": total_records,
        "total_dss": total_dss,
        "block_count": block_count,
        "zero_forecast_plants": full_zero_unique_dss,
        "missing_block_plants": len(missing_blocks),
        "best_nrmse": best,
        "worst_nrmse": worst,
        "net_bias_mw": bias.get("net_bias_mw", 0),
    }


# ============================================================
# EXPORT REPORT
# ============================================================

def make_dashboard_data(report_date, company, revision):
    conn = get_conn()

    with conn.cursor() as cur:
        summary = get_forecast_summary(cur, report_date, company, revision)
        forecast_curve = get_forecast_curve(cur, report_date, company, revision)
        actual_curve = get_actual_curve(report_date)

        daily_nrmse = get_daily_nrmse(cur, report_date, company, revision)
        monthly_nrmse = get_monthly_nrmse(cur, report_date, company, revision)

        revision_summary = get_revision_summary(forecast_curve)
        bias = get_bias_analysis(forecast_curve, actual_curve)

        plant_errors = get_plant_forecast_ranking(cur, report_date, company, revision)
        zero_plants = get_zero_forecast_plants(cur, report_date, company, revision)
        missing_blocks = get_missing_blocks(cur, report_date, company, revision)

        data_availability = get_data_availability(cur, report_date)
        scada_status = get_scada_data_status(report_date)
        payment_impact = get_payment_impact(daily_nrmse)

        solar_zero_blockwise = get_solar_zero_forecast_blockwise(
            cur, report_date, company, revision
        )
        wind_zero_blockwise = get_wind_zero_forecast_blockwise(
            cur, report_date, company, revision
        )

        repeated_forecast = get_repeated_forecast_count(
            cur, report_date, company, revision
        )

        solar_scada_night = get_solar_scada_night_alert(report_date)
        model_status = get_forecast_model_status(report_date, revision)

        alerts = build_alerts(
            daily_nrmse=daily_nrmse,
            zero_plants=zero_plants,
            missing_blocks=missing_blocks,
            data_availability=data_availability,
            solar_zero_blockwise=solar_zero_blockwise,
            wind_zero_blockwise=wind_zero_blockwise,
            repeated_forecast=repeated_forecast,
            solar_scada_night=solar_scada_night,
            model_status=model_status,
            scada_status=scada_status,
        )

    conn.close()

    return {
        "summary": summary,
        "forecast_curve": forecast_curve,
        "actual_curve": actual_curve,
        "daily_nrmse": daily_nrmse,
        "monthly_nrmse": monthly_nrmse,
        "revision_summary": revision_summary,
        "bias": bias,
        "plant_errors": plant_errors,
        "zero_plants": zero_plants,
        "missing_blocks": missing_blocks,
        "data_availability": data_availability,
        "scada_status": scada_status,
        "payment_impact": payment_impact,
        "alerts": alerts,
        "model_status": model_status,
        "solar_zero_blockwise": solar_zero_blockwise,
        "wind_zero_blockwise": wind_zero_blockwise,
    }


@router.get("/api/export/{export_type}")
def export_report(
    export_type: str,
    report_date: str = Query(default=None),
    company: str = Query(default="ALL"),
    revision: str = Query(default="ALL"),
):
    if not report_date:
        report_date = str(date.today())

    data = make_dashboard_data(report_date, company, revision)

    if export_type == "csv":
        return export_csv(report_date, company, revision, data)

    if export_type == "excel":
        return export_excel(report_date, company, revision, data)

    if export_type == "pdf":
        return export_pdf(report_date, company, revision, data)

    return JSONResponse(
        status_code=400,
        content={
            "error": True,
            "message": "Invalid export type. Use csv, excel, or pdf.",
        },
    )


def export_csv(report_date, company, revision, data):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Daily Forecast Analysis Report"])
    writer.writerow(["Date", report_date])
    writer.writerow(["Company", company])
    writer.writerow(["Revision", revision])
    writer.writerow([])

    writer.writerow(["Forecast Model Status"])
    writer.writerow(["Status", data["model_status"].get("status")])
    writer.writerow(["Message", data["model_status"].get("message")])
    writer.writerow([])

    writer.writerow(["SCADA Status"])
    writer.writerow(["Status", data["scada_status"].get("status")])
    writer.writerow(["Rows", data["scada_status"].get("rows")])
    writer.writerow(["DSS Count", data["scada_status"].get("dss_count")])
    writer.writerow(["From", data["scada_status"].get("from_time")])
    writer.writerow(["To", data["scada_status"].get("to_time")])
    writer.writerow(["Total SCADA", data["scada_status"].get("total_scada")])
    writer.writerow([])

    writer.writerow(["Alerts"])
    for a in data["alerts"]:
        writer.writerow([a])
    writer.writerow([])

    writer.writerow(["Daily NRMSE"])
    writer.writerow(["Date", "Company", "Energy", "Revision", "NRMSE", "Status"])
    for r in data["daily_nrmse"]:
        writer.writerow([
            r.get("date"),
            r.get("company"),
            r.get("energy_type"),
            r.get("revision_num"),
            r.get("nrmse_values"),
            r.get("status"),
        ])

    writer.writerow([])
    writer.writerow(["Forecast Summary"])
    writer.writerow(["Company", "Revision", "DSS", "Blocks", "Total MW", "Zero Rows"])
    for r in data["summary"]:
        writer.writerow([
            r.get("company"),
            r.get("revision_number"),
            r.get("dss_count"),
            r.get("block_count"),
            r.get("total_forecast_mw"),
            r.get("zero_rows"),
        ])

    output.seek(0)

    filename = f"forecast_report_{report_date}_{company}_{revision}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )


def export_excel(report_date, company, revision, data):
    wb = Workbook()

    ws = wb.active
    ws.title = "Report Summary"

    ws.append(["Daily Forecast Analysis Report"])
    ws.append(["Date", report_date])
    ws.append(["Company", company])
    ws.append(["Revision", revision])
    ws.append([])

    ws.append(["Forecast Model Status"])
    ws.append(["Status", data["model_status"].get("status")])
    ws.append(["Message", data["model_status"].get("message")])
    ws.append([])

    ws.append(["SCADA Status"])
    ws.append(["Status", data["scada_status"].get("status")])
    ws.append(["Rows", data["scada_status"].get("rows")])
    ws.append(["DSS Count", data["scada_status"].get("dss_count")])
    ws.append(["From", data["scada_status"].get("from_time")])
    ws.append(["To", data["scada_status"].get("to_time")])
    ws.append(["Total SCADA", data["scada_status"].get("total_scada")])
    ws.append([])

    ws.append(["Alerts"])
    for a in data["alerts"]:
        ws.append([a])

    ws2 = wb.create_sheet("Daily NRMSE")
    ws2.append(["Date", "Company", "Energy", "Revision", "NRMSE", "Status"])
    for r in data["daily_nrmse"]:
        ws2.append([
            str(r.get("date")),
            r.get("company"),
            r.get("energy_type"),
            r.get("revision_num"),
            r.get("nrmse_values"),
            r.get("status"),
        ])

    ws3 = wb.create_sheet("Forecast Summary")
    ws3.append(["Company", "Revision", "DSS", "Blocks", "Total MW", "Zero Rows"])
    for r in data["summary"]:
        ws3.append([
            r.get("company"),
            r.get("revision_number"),
            r.get("dss_count"),
            r.get("block_count"),
            r.get("total_forecast_mw"),
            r.get("zero_rows"),
        ])

    ws4 = wb.create_sheet("Payment Impact")
    ws4.append(["Company", "Energy", "Revision", "NRMSE", "Pay", "Lost"])
    for r in data["payment_impact"]:
        ws4.append([
            r.get("company"),
            r.get("energy_type"),
            r.get("revision_num"),
            r.get("nrmse_values"),
            r.get("eligible_pay"),
            r.get("lost_pay"),
        ])

    ws5 = wb.create_sheet("Full Zero Forecast")
    ws5.append(["Company", "DSS ID", "Revision", "Total Blocks", "Zero Blocks", "Status"])
    for r in data["zero_plants"]:
        ws5.append([
            r.get("company"),
            r.get("dss_id"),
            r.get("revision_number"),
            r.get("total_blocks"),
            r.get("zero_blocks"),
            r.get("status"),
        ])

    ws6 = wb.create_sheet("Solar Zero Blockwise")
    ws6.append(["Block No", "Block Time", "Zero Plants"])
    for r in data.get("solar_zero_blockwise", []):
        ws6.append([r.get("block_no"), r.get("block_time"), r.get("zero_plants")])

    ws7 = wb.create_sheet("Wind Zero Blockwise")
    ws7.append(["Block No", "Block Time", "Zero Plants"])
    for r in data.get("wind_zero_blockwise", []):
        ws7.append([r.get("block_no"), r.get("block_time"), r.get("zero_plants")])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    filename = f"forecast_report_{report_date}_{company}_{revision}.xlsx"

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


def export_pdf(report_date, company, revision, data):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")

    doc = SimpleDocTemplate(
        tmp.name,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Daily Forecast Analysis Report", styles["Title"]))
    elements.append(Paragraph(f"Date: {report_date}", styles["Normal"]))
    elements.append(Paragraph(f"Company: {company}", styles["Normal"]))
    elements.append(Paragraph(f"Revision: {revision}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Forecast Model Status", styles["Heading2"]))
    elements.append(Paragraph(data["model_status"].get("message", "-"), styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("SCADA Status", styles["Heading2"]))
    elements.append(Paragraph(
        f"Status: {data['scada_status'].get('status')} | "
        f"Rows: {data['scada_status'].get('rows')} | "
        f"DSS Count: {data['scada_status'].get('dss_count')} | "
        f"From: {data['scada_status'].get('from_time')} | "
        f"To: {data['scada_status'].get('to_time')} | "
        f"Total SCADA: {data['scada_status'].get('total_scada')} MW",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Alerts", styles["Heading2"]))

    alert_table = [["No", "Alert"]]

    for idx, a in enumerate(data["alerts"], start=1):
        alert_table.append([idx, a])

    t = Table(alert_table, colWidths=[40, 700])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    elements.append(t)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Daily NRMSE", styles["Heading2"]))

    nrmse_table = [["Company", "Energy", "Revision", "NRMSE", "Status"]]

    for r in data["daily_nrmse"]:
        nrmse_table.append([
            r.get("company"),
            r.get("energy_type"),
            r.get("revision_num"),
            r.get("nrmse_values"),
            r.get("status"),
        ])

    t2 = Table(nrmse_table)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(t2)

    doc.build(elements)

    filename = f"forecast_report_{report_date}_{company}_{revision}.pdf"

    return FileResponse(
        tmp.name,
        media_type="application/pdf",
        filename=filename,
    )