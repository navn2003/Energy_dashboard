"""
upload_kptcl.py — Parse the KPTCL Excel file and load into power_market_db.

Handles target areas:
  1. `unitmaster`  sheet  → unit_master table
  2. `cost`        sheet  → generation_cost_master table
  3. 6 wide sheets (URS, SCH, ENT, Backdown, Entitlement, Calculated-DC) → plant_block_data
  4. 4 calc sheets (MustRun, Minimum_Capacity, etc) → calculated_values
"""
import hashlib
from datetime import datetime, date, time, timedelta
from io import BytesIO
from collections import defaultdict

import openpyxl
import pymysql


# ======================= HELPERS =======================

def _num(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def _int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def _date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("true", "1", "yes"):
        return 1
    if s in ("false", "0", "no"):
        return 0
    return None

def _row_hash(fields: dict) -> str:
    parts = [f"{k}={'' if v is None else v}" for k, v in sorted(fields.items())]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()

def _find_header_row(ws, marker="date"):
    for i, row in enumerate(ws.iter_rows(max_col=2, values_only=True), 1):
        if row[0] is not None and str(row[0]).strip().lower() == marker:
            return i
    return None

def _block_time(block_no):
    mins = (block_no - 1) * 15
    return time(mins // 60, mins % 60)


# ======================= UNIT MASTER =======================

def _parse_unitmaster(wb):
    ws = wb["unitmaster"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        un = _str(r[0])
        if not un: continue
        rows.append({
            "unit_name": un,
            "generator": _str(r[1]),
            "plant_type": _str(r[2]),
            "fuel_type": _str(r[3]),
            "total_capacity": _num(r[4]),
            "contracted_capacity": _num(r[5]),
            "min_technical_limit": _num(r[6]),
            "time_between_ramp_up_down": _num(r[7]),
            "ramp_rate_mw_per_min": _num(r[8]),
            "pool": _str(r[9]),
            "must_run": _bool(r[10]),
            "variable_cost": _num(r[11]),
        })
    return rows

def _load_unitmaster(conn, rows, source_file):
    cur = conn.cursor()
    inserted = 0
    for r in rows:
        rh = _row_hash(r)
        try:
            cur.execute(
                """INSERT INTO unit_master
                     (unit_name, generator, plant_type, fuel_type,
                      total_capacity, contracted_capacity, min_technical_limit,
                      time_between_ramp_up_down, ramp_rate_mw_per_min,
                      pool, must_run, variable_cost, row_hash, source_file)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   ON DUPLICATE KEY UPDATE id=id""",
                (r["unit_name"], r["generator"], r["plant_type"], r["fuel_type"],
                 r["total_capacity"], r["contracted_capacity"], r["min_technical_limit"],
                 r["time_between_ramp_up_down"], r["ramp_rate_mw_per_min"],
                 r["pool"], r["must_run"], r["variable_cost"], rh, source_file)
            )
            if cur.rowcount == 1: inserted += 1
        except pymysql.err.IntegrityError:
            pass 
    conn.commit()
    cur.close()
    return inserted


# ======================= GENERATION COST MASTER =======================

def _parse_cost(wb):
    ws = wb["cost"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        un = _str(r[0])
        if not un: continue
        rows.append({
            "unit_name": un, "generator": _str(r[1]), "station_name": _str(r[2]), "acronym": _str(r[3]),
            "contracted_capacity": _num(r[4]), "valid_from": _date(r[5]), "valid_to": _date(r[6]),
            "fixed_cost": _num(r[7]), "variable_cost": _num(r[8]), "must_run": _bool(r[9]),
            "mini_limit": _num(r[10]), "pool": _str(r[11]), "plant_type": _str(r[12]),
            "entire_shared_ent": _num(r[13]), "entitlement_per_unit": _num(r[14]),
            "no_of_units": _int(r[15]), "mintech_per_unit": _num(r[16]), "mini_limit_percent": _num(r[17]),
        })
    return rows

def _load_cost(conn, rows, source_file):
    cur = conn.cursor()
    inserted = 0
    for r in rows:
        rh = _row_hash(r)
        try:
            cur.execute(
                """INSERT INTO generation_cost_master
                     (unit_name, generator, station_name, acronym,
                      contracted_capacity, valid_from, valid_to,
                      fixed_cost, variable_cost, must_run, mini_limit,
                      pool, plant_type, entire_shared_ent,
                      entitlement_per_unit, no_of_units,
                      mintech_per_unit, mini_limit_percent,
                      row_hash, source_file)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   ON DUPLICATE KEY UPDATE id=id""",
                (r["unit_name"], r["generator"], r["station_name"], r["acronym"],
                 r["contracted_capacity"], r["valid_from"], r["valid_to"],
                 r["fixed_cost"], r["variable_cost"], r["must_run"], r["mini_limit"],
                 r["pool"], r["plant_type"], r["entire_shared_ent"],
                 r["entitlement_per_unit"], r["no_of_units"],
                 r["mintech_per_unit"], r["mini_limit_percent"],
                 rh, source_file)
            )
            if cur.rowcount == 1: inserted += 1
        except pymysql.err.IntegrityError:
            pass
    conn.commit()
    cur.close()
    return inserted


# ======================= SHARED BLOCK KEYS =======================

def _ensure_block_keys(conn, dates_blocks):
    cur = conn.cursor()
    cur.execute("SELECT block_id, report_date, block_no FROM plant_block_key")
    existing = {}
    for r in cur.fetchall():
        existing[(r[1], r[2])] = r[0]

    cur.execute("SELECT COALESCE(MAX(block_id), 999) FROM plant_block_key")
    next_id = cur.fetchone()[0] + 1

    to_insert = []
    for (d, b) in sorted(dates_blocks):
        if (d, b) not in existing:
            bt = _block_time(b)
            to_insert.append((next_id, d, b, bt))
            existing[(d, b)] = next_id
            next_id += 1

    if to_insert:
        cur.executemany(
            "INSERT IGNORE INTO plant_block_key (block_id, report_date, block_no, block_time) VALUES (%s,%s,%s,%s)",
            to_insert,
        )
        conn.commit()
    cur.close()
    return existing

def _resolve_unit_ids(conn):
    cur = conn.cursor()
    cur.execute("SELECT unit_name, id FROM unit_master")
    mapping = {str(r[0]).strip().upper(): r[1] for r in cur.fetchall()}
    cur.close()
    return mapping


# ======================= PLANT BLOCK DATA (6 wide sheets) =======================

_WIDE_SHEETS = {
    "URS":           "urs",
    "SCH":           "sch",
    "ENT":           "ent",
    "Backdown":      "backdown",
    "Entitlement":   "entitlement",
    "Calculated-DC": "calculated_dc",
}

def _parse_wide_sheets(wb):
    data = defaultdict(dict)
    for sheet_name, field in _WIDE_SHEETS.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        hrow = _find_header_row(ws, "date")
        if hrow is None: continue
        header = list(ws.iter_rows(min_row=hrow, max_row=hrow, values_only=True))[0]
        units = []
        for i, h in enumerate(header):
            if i < 2: continue
            un = _str(h)
            if un: units.append((i, un))
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            d = _date(row[0])
            b = _int(row[1])
            if d is None or b is None: continue
            for col_idx, un in units:
                if col_idx >= len(row): continue
                val = _num(row[col_idx])
                data[(d, b, un)][field] = val
    return data

def _load_plant_block_data(conn, data, source_file):
    cur = conn.cursor()
    dates_blocks = set((d, b) for (d, b, _) in data.keys())
    bk_map = _ensure_block_keys(conn, dates_blocks)
    
    # We fetch unit_master ID to insert into plant_block_data
    uid_map = _resolve_unit_ids(conn)

    cur.execute("SELECT timestamp_id, unit_name FROM plant_block_data")
    existing = set((r[0], r[1]) for r in cur.fetchall())

    to_insert = []
    skipped = 0
    unresolved_units = set()
    for (d, b, un), vals in data.items():
        ts_id = bk_map.get((d, b))
        if ts_id is None: continue
        u_id = uid_map.get(un.upper())
        if u_id is None:
            unresolved_units.add(un)

        if (ts_id, un) in existing:
            skipped += 1
            continue

        to_insert.append((
            ts_id, u_id, un,
            vals.get("urs"), vals.get("sch"), vals.get("ent"),
            vals.get("backdown"), vals.get("entitlement"),
            vals.get("calculated_dc"), source_file,
        ))

    inserted = 0
    if to_insert:
        cur.executemany(
            """INSERT INTO plant_block_data
                 (timestamp_id, unit_id, unit_name,
                  urs, sch, ent, backdown, entitlement,
                  calculated_dc, source_file)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            to_insert,
        )
        inserted = cur.rowcount
        conn.commit()
    cur.close()
    return inserted, skipped, unresolved_units


# ======================= CALCULATED VALUES (4 wide sheets) =======================

def _parse_calc_sheets(wb):
    data = defaultdict(dict)
    for sheet_name in wb.sheetnames:
        field = None
        if sheet_name == "MustRun": field = "must_run_value"
        elif sheet_name == "Minimum_Capacity": field = "minimum_capacity"
        elif sheet_name == "CapacityOverMin": field = "capacity_over_min"
        elif sheet_name.startswith("CapOverMin") or sheet_name.startswith("CapacityOverMinAct"): field = "capacity_over_min_act_sch"
        
        if not field: continue

        ws = wb[sheet_name]
        hrow = _find_header_row(ws, "date")
        if hrow is None: continue
        header = list(ws.iter_rows(min_row=hrow, max_row=hrow, values_only=True))[0]
        units = []
        for i, h in enumerate(header):
            if i < 2: continue
            un = _str(h)
            if un: units.append((i, un))
            
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            d = _date(row[0])
            b = _int(row[1])
            if d is None or b is None: continue
            for col_idx, un in units:
                if col_idx >= len(row): continue
                val = _num(row[col_idx])
                data[(d, b, un)][field] = val
    return data

def _load_calculated_values(conn, data):
    cur = conn.cursor()
    dates_blocks = set((d, b) for (d, b, _) in data.keys())
    bk_map = _ensure_block_keys(conn, dates_blocks)
    uid_map = _resolve_unit_ids(conn)

    to_insert = []
    for (d, b, un), vals in data.items():
        ts_id = bk_map.get((d, b))
        u_id = uid_map.get(un.upper())
        if ts_id is None or u_id is None:
            continue
            
        to_insert.append((
            ts_id, u_id,
            vals.get("must_run_value", 0.0),
            vals.get("minimum_capacity", 0.0),
            vals.get("capacity_over_min", 0.0),
            vals.get("capacity_over_min_act_sch", 0.0)
        ))

    inserted = 0
    if to_insert:
        # Delete existing data for these timestamp_id & unit_id combos to prevent duplicates before inserting
        for ts_id, u_id, *_ in to_insert:
            cur.execute("DELETE FROM calculated_values WHERE timestamp_id=%s AND unit_id=%s", (ts_id, u_id))
        
        cur.executemany(
            """INSERT INTO calculated_values
                 (timestamp_id, unit_id, must_run_value, minimum_capacity, capacity_over_min, capacity_over_min_act_sch)
               VALUES (%s,%s,%s,%s,%s,%s)
            """,
            to_insert
        )
        inserted = cur.rowcount
        conn.commit()
    cur.close()
    return inserted


# ======================= MAIN ENTRY POINT =======================

def process_upload(file_bytes: bytes, filename: str, db_config: dict):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    conn = pymysql.connect(**db_config, charset="utf8mb4", autocommit=False)
    conn.cursor().execute("USE power_market_db")

    summary = {"filename": filename, "tables": {}}
    errors = []

    try:
        # 1. Unit Master
        if "unitmaster" in wb.sheetnames:
            try:
                rows = _parse_unitmaster(wb)
                ins = _load_unitmaster(conn, rows, filename)
                summary["tables"]["unit_master"] = {"parsed": len(rows), "inserted": ins, "skipped": len(rows) - ins}
            except Exception as e:
                errors.append(f"unitmaster: {e}")

        # 2. Cost
        if "cost" in wb.sheetnames:
            try:
                rows = _parse_cost(wb)
                ins = _load_cost(conn, rows, filename)
                summary["tables"]["generation_cost_master"] = {"parsed": len(rows), "inserted": ins, "skipped": len(rows) - ins}
            except Exception as e:
                errors.append(f"cost: {e}")

        # 3. Plant Block Data (6 sheets)
        try:
            data = _parse_wide_sheets(wb)
            if data:
                ins, skip, unresolved = _load_plant_block_data(conn, data, filename)
                summary["tables"]["plant_block_data"] = {
                    "parsed": len(data),
                    "inserted": ins,
                    "skipped": skip,
                    "unresolved_units": sorted(unresolved) if unresolved else [],
                }
            else:
                summary["tables"]["plant_block_data"] = {"parsed": 0, "inserted": 0, "skipped": 0}
        except Exception as e:
            errors.append(f"plant_block_data: {e}")

        # 4. Calculated Values (4 sheets)
        try:
            calc_data = _parse_calc_sheets(wb)
            if calc_data:
                ins = _load_calculated_values(conn, calc_data)
                summary["tables"]["calculated_values"] = {
                    "parsed": len(calc_data),
                    "inserted": ins,
                    "skipped": 0
                }
            else:
                 summary["tables"]["calculated_values"] = {"parsed": 0, "inserted": 0, "skipped": 0}
        except Exception as e:
            errors.append(f"calculated_values: {e}")

    finally:
        conn.close()
        wb.close()

    summary["errors"] = errors
    summary["status"] = "success" if not errors else "partial"
    return summary