from fastapi import FastAPI, Query, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse, Response
from database import get_connection
from datetime import date, timedelta
import calendar
import pandas as pd
import os
import re
import pymysql
import subprocess
import uuid
from pathlib import Path
from typing import Optional, Dict, Any
from pydantic import BaseModel
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

app = FastAPI(title="Energy Monitor API")
# ... (rest of your code)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
# ── Authentication ───────────────────────────────────────────────────────────
import os, hmac, hashlib, base64, time
from fastapi import Request
from pydantic import BaseModel

# Credentials & secret — OVERRIDE THESE VIA ENV (esp. on public Replit!).
AUTH_USER   = os.getenv("AUTH_USER", "Admin")
AUTH_PASS   = os.getenv("AUTH_PASS", "Energy@123")
AUTH_SECRET = os.getenv("AUTH_SECRET", "please-change-this-long-random-secret-7f3a9c21")
AUTH_TTL    = int(os.getenv("AUTH_TTL", "43200"))   # token lifetime in seconds (12h)

# /api/ paths that do NOT require a token:
_PUBLIC_API = {"/api/login"}


def _make_token(username: str) -> str:
    exp = int(time.time()) + AUTH_TTL
    msg = f"{username}:{exp}"
    sig = hmac.new(AUTH_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(f"{msg}:{sig}".encode()).decode()


def _verify_token(token: str):
    try:
        raw = base64.urlsafe_b64decode(token.encode()).decode()
        username, exp, sig = raw.rsplit(":", 2)
        if int(exp) < time.time():
            return None
        expected = hmac.new(AUTH_SECRET.encode(), f"{username}:{exp}".encode(),
                            hashlib.sha256).hexdigest()
        return username if hmac.compare_digest(sig, expected) else None
    except Exception:
        return None


class _LoginIn(BaseModel):
    username: str
    password: str


@app.post("/api/login")
def api_login(body: _LoginIn):
    if hmac.compare_digest(body.username, AUTH_USER) and hmac.compare_digest(body.password, AUTH_PASS):
        return {"token": _make_token(body.username), "username": body.username}
    return JSONResponse({"error": "Invalid username or password."}, status_code=401)


@app.middleware("http")
async def _auth_guard(request: Request, call_next):
    path = request.url.path
    if path.startswith("/api/") and path not in _PUBLIC_API and request.method != "OPTIONS":
        auth = request.headers.get("Authorization", "")
        token = auth[7:] if auth.startswith("Bearer ") else ""
        if not token:
            token = request.query_params.get("token", "")
        if not _verify_token(token):
            return JSONResponse({"error": "Unauthorized", "auth_required": True},
                                status_code=401)
    return await call_next(request)
# ── NRMSE endpoints ──────────────────────────────────────────────────────────

@app.get("/api/nrmse/daily")
def get_nrmse_daily(date: str = Query(..., description="YYYY-MM-DD")):
    """Returns NRMSE values for a single day grouped by company, energy_type, revision_num."""
    sql = """
        SELECT company, energy_type, revision_num, nrmse_values
        FROM daily_nrmse_values
        WHERE date = %s
        ORDER BY energy_type, company, revision_num
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, (date,))
    rows = cursor.fetchall()

    result = _build_nrmse_response(rows)

    try:
        cursor.execute(
            "SELECT energy_type, nrmse FROM quenext_nrmse WHERE `date` = %s",
            (date,),
        )
        qn = {}
        for r in cursor.fetchall():
            et = (r["energy_type"] or "").strip().lower()
            qn[et] = float(r["nrmse"]) if r["nrmse"] is not None else None

        for et in ("solar", "wind"):
            if et not in result:
                result[et] = {}
            qkey = next((k for k in result[et].keys() if k.lower() == "quenext"), "Quenext")
            result[et].setdefault(qkey, {})
            result[et][qkey]["R16"] = qn.get(et, None)
    except Exception as e:
        print(f"quenext_nrmse override error: {e}")

    cursor.close()
    conn.close()
    return result


@app.get("/api/nrmse/monthly")
def get_nrmse_monthly(year: int = Query(...), month: int = Query(...)):
    """Returns precomputed monthly NRMSE values from monthly_nrmse_values."""
    sql = """
        SELECT company, energy_type, revision_num, nrmse_values
        FROM monthly_nrmse_values
        WHERE YEAR(`month`) = %s AND MONTH(`month`) = %s
        ORDER BY energy_type, company, revision_num
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, (year, month))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return _build_monthly_nrmse_response(rows)


def _build_monthly_nrmse_response(rows):
    MONTHLY_REV_MAP = {
        "R20": "R16",
        "R02": "DA2",
        "R-16": "R16",
        "RD2":  "DA2",
    }

    result = {"solar": {}, "wind": {}}

    for row in rows:
        company     = row["company"]
        energy_type = row["energy_type"].lower()
        rev_raw     = row["revision_num"]
        nrmse_val   = float(row["nrmse_values"]) if row["nrmse_values"] is not None else None
        rev_label   = MONTHLY_REV_MAP.get(rev_raw, rev_raw)

        if energy_type not in result:
            result[energy_type] = {}
        if company not in result[energy_type]:
            result[energy_type][company] = {}

        result[energy_type][company][rev_label] = nrmse_val

    return result

@app.get("/api/nrmse/custom")
def get_nrmse_custom(
    from_date: str = Query(..., description="YYYY-MM-DD"),
    to_date: str = Query(..., description="YYYY-MM-DD"),
):
    """Returns average NRMSE values for a custom date range."""
    sql = """
        SELECT company, energy_type, revision_num,
        AVG(nrmse_values) AS nrmse_values
        FROM daily_nrmse_values
        WHERE date BETWEEN %s AND %s
        GROUP BY company, energy_type, revision_num
        ORDER BY energy_type, company, revision_num
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, (from_date, to_date))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return _build_nrmse_response(rows)


def _build_nrmse_response(rows):
    REV_MAP = {
        "R20": "R16",
        "R02": "DA2",
        "R-16": "R16",
        "RD2":  "DA2",
    }

    result = {
        "solar": {},
        "wind": {},
    }

    for row in rows:
        company     = row["company"]
        energy_type = row["energy_type"].lower()
        rev_raw     = row["revision_num"]
        nrmse_val   = float(row["nrmse_values"]) if row["nrmse_values"] is not None else None
        rev_label   = REV_MAP.get(rev_raw, rev_raw)   # falls back to raw value if unmapped

        if energy_type not in result:
            result[energy_type] = {}
        if company not in result[energy_type]:
            result[energy_type][company] = {}

        result[energy_type][company][rev_label] = nrmse_val

    return result


# ── NRMSE Monitor (per-day green/yellow/red breakdown) ────────────────────────

_MONITOR_REV_MAP = {
    "R20": "R16",   # ISPL intra-day
    "R-16": "R16",  # Quenext intra-day
    "R02": "R02",   # ISPL day-ahead
    "RD2": "R02",   # Quenext day-ahead
    "DA2": "R02",
    "R16": "R16",
}

_MONITOR_THRESHOLDS = {
    "SOLAR": {"R16": [3, 5],  "R02": [5, 10]},
    "WIND":  {"R16": [5, 10], "R02": [10, 15]},
}

def _classify(value, lo, hi):
    """green if <=lo, yellow if <=hi, else red. None -> None."""
    if value is None:
        return None
    v = float(value)
    if v <= lo:
        return "green"
    if v <= hi:
        return "yellow"
    return "red"

def _monitor_payload(year: int, month: int, energy_type: str):
    energy = (energy_type or "SOLAR").upper()
    thresholds = _MONITOR_THRESHOLDS.get(energy, _MONITOR_THRESHOLDS["SOLAR"])

    sql = """
        SELECT DAY(date) AS day, company, revision_num, nrmse_values
        FROM daily_nrmse_values
        WHERE YEAR(date) = %s AND MONTH(date) = %s AND UPPER(energy_type) = %s
        ORDER BY day
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(sql, (year, month, energy))
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    days_in_month = calendar.monthrange(year, month)[1]

    def blank_company():
        days = {
            d: {"R16": {"value": None, "status": None},
                "R02": {"value": None, "status": None}}
            for d in range(1, days_in_month + 1)
        }
        counts = {"R16": {"green": 0, "yellow": 0, "red": 0},
                  "R02": {"green": 0, "yellow": 0, "red": 0}}
        return {"days": days, "counts": counts}

    companies: Dict[str, Any] = {"ISPL": blank_company(), "Quenext": blank_company()}

    for row in rows:
        raw_company = str(row["company"] or "").strip().upper()
        if "QUENEXT" in raw_company:
            comp = "Quenext"
        elif "ISPL" in raw_company:
            comp = "ISPL"
        else:
            continue

        rev = _MONITOR_REV_MAP.get(str(row["revision_num"]).strip())
        if rev not in ("R16", "R02"):
            continue

        day = int(row["day"])
        val = float(row["nrmse_values"]) if row["nrmse_values"] is not None else None
        lo, hi = thresholds[rev]
        status = _classify(val, lo, hi)

        cell = companies[comp]["days"].get(day)
        if cell is None:
            continue
        cell[rev] = {"value": round(val, 4) if val is not None else None, "status": status}
        if status:
            companies[comp]["counts"][rev][status] += 1

    for comp in companies:
        day_map = companies[comp]["days"]
        companies[comp]["days"] = [
            {"date": d, "R16": day_map[d]["R16"], "R02": day_map[d]["R02"]}
            for d in range(1, days_in_month + 1)
        ]

    return {
        "year": year,
        "month": month,
        "energy_type": energy,
        "days_in_month": days_in_month,
        "thresholds": thresholds,
        "companies": companies,
    }


@app.get("/api/nrmse/monitor")
def get_nrmse_monitor(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    energy_type: str = Query("SOLAR", description="SOLAR or WIND"),
):
    """Per-day green/yellow/red classification for ISPL & Quenext, revisions R16 & R02."""
    return _monitor_payload(year, month, energy_type)


@app.get("/api/nrmse/monitor-csv")
def get_nrmse_monitor_csv(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    energy_type: str = Query("SOLAR"),
):
    """CSV export of the per-day NRMSE classification."""
    data = _monitor_payload(year, month, energy_type)

    lines = ["Company,Date,R16_Value,R16_Status,R02_Value,R02_Status"]
    for comp, cdata in data["companies"].items():
        for d in cdata["days"]:
            r16, r02 = d["R16"], d["R02"]
            date_str = f"{year:04d}-{month:02d}-{d['date']:02d}"
            lines.append(",".join([
                comp,
                date_str,
                "" if r16["value"] is None else str(r16["value"]),
                r16["status"] or "",
                "" if r02["value"] is None else str(r02["value"]),
                r02["status"] or "",
            ]))

    csv_text = "\n".join(lines)
    fname = f"nrmse_monitor_{data['energy_type']}_{year}_{month:02d}.csv"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Plant endpoints ──────────────────────────────────────────────────────────

@app.get("/api/plants")
def get_plants():
    """Returns all plants from plant_master."""
    sql = """
        SELECT plant_id, plant_name, plant_type, capacity_mw
        FROM plant_master
        ORDER BY plant_type, plant_name
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(sql)
        rows = cursor.fetchall()
    except Exception as e:
        cursor.close()
        conn.close()
        return {"error": str(e), "plants": []}
    cursor.close()
    conn.close()
    return {"plants": rows}


# ── Chart data (96 slots) ─────────────────────────────────────────────────────

@app.get("/api/chart/daily")
def get_chart_daily(
    date: str = Query(..., description="YYYY-MM-DD"),
    plant_id: str = Query(None, description="Optional plant filter"),
    energy_type: str = Query(None, description="SOLAR or WIND"),
    revision: str = Query("R16", description="Logical revision: R16 (intra-day) or DA2 (day-ahead)"),
):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    slots = _generate_slots()
    result_map = {slot: {"slot": slot, "actual": None, "ISPL": None, "Quenext": None} for slot in slots}

    rev_logical = (revision or "R16").upper().replace("-", "").replace(" ", "")
    REVISION_MAP = {
        "R16": {"ISPL": "R20",  "Quenext": "R-16"},
        "DA2": {"ISPL": "R02",  "Quenext": "RD2"},
    }
    vendor_codes = REVISION_MAP.get(rev_logical, REVISION_MAP["R16"])
    energy = (energy_type or "").upper()

    def _norm(code):
        return code.upper().replace("-", "").replace(" ", "")

    fc_sql = """
        SELECT DATE_FORMAT(f.timestamp, '%H:%i') AS time_slot,
               SUM(f.forecast_mw) AS mw
        FROM {table} f
        JOIN plant_master p ON p.dss_id = f.dss_id
        WHERE DATE(f.timestamp) = %s
          AND ( %s = '' OR UPPER(p.plant_type) = %s )
          AND REPLACE(REPLACE(UPPER(f.revision_number), '-', ''), ' ', '') = %s
        GROUP BY time_slot
    """

    def _load_vendor(table, raw_rev, out_key):
        try:
            cursor.execute(fc_sql.format(table=table),
                           (date, energy, energy, _norm(raw_rev)))
            for r in cursor.fetchall():
                slot = r["time_slot"]
                if slot in result_map and r["mw"] is not None:
                    result_map[slot][out_key] = float(r["mw"])
        except Exception as e:
            print(f"Chart {out_key} query error: {e}")

    _load_vendor("dss_forecast_ispl", vendor_codes["ISPL"], "ISPL")

    if rev_logical == "R16":
        comp_sql = """
            SELECT DATE_FORMAT(timestamp, '%H:%i') AS time_slot,
            SUM(total_forecast_mw) AS mw
            FROM composite_r05_r20_forecast_96
            WHERE DATE(timestamp) = %s
              AND ( %s = '' OR UPPER(energy_type) = %s )
              AND UPPER(company_name) = 'QUENEXT'
            GROUP BY time_slot
        """
        try:
            cursor.execute(comp_sql, (date, energy, energy))
            for r in cursor.fetchall():
                slot = r["time_slot"]
                if slot in result_map and r["mw"] is not None:
                    result_map[slot]["Quenext"] = float(r["mw"])
        except Exception as e:
            print(f"Chart Quenext (composite) query error: {e}")
    else:
        _load_vendor("dss_forecast_quenext", vendor_codes["Quenext"], "Quenext")

    actual_sql = """
        SELECT DATE_FORMAT(timestamp, '%H:%i') AS time_slot,
               MAX(total_scada) AS actual
        FROM dss_forecast_aggregated
        WHERE DATE(timestamp) = %s
          AND UPPER(energy_type) = %s
        GROUP BY time_slot
    """
    try:
        cursor.execute(actual_sql, (date, energy))
        for r in cursor.fetchall():
            slot = r["time_slot"]
            if slot in result_map and r["actual"] is not None:
                result_map[slot]["actual"] = float(r["actual"])
    except Exception as e:
        print(f"Chart actual query error: {e}")
    finally:
        cursor.close()
        conn.close()

    return {"slots": list(result_map.values())}

def _generate_slots():
    """Generates 96 time slots: 00:00, 00:15, ... 23:45"""
    slots = []
    for h in range(24):
        for m in (0, 15, 30, 45):
            slots.append(f"{h:02d}:{m:02d}")
    return slots

# ── Available dates ───────────────────────────────────────────────────────────

@app.get("/api/available-dates")
def get_available_dates():
    sql = "SELECT DISTINCT date FROM daily_nrmse_values ORDER BY date DESC LIMIT 90"
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return {"dates": [str(r[0]) for r in rows]}

@app.get("/api/health")
def health():
    return {"status": "ok"}


# =============================================================================
# ── SCADA REPORTS INTEGRATION ────────────────────────────────────────────────
# =============================================================================

SCADA_DB_CONFIG = {
    "host": "13.205.184.74",
    "user": "normal_access",
    "password": "energyX@123#",
    "database": "engxai_fs",
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}

def get_scada_connection():
    return pymysql.connect(**SCADA_DB_CONFIG)

def clean_dss_id(value):
    if value is None: return None
    value = str(value).strip()
    if value == "" or value.upper() == "DSS_ID": return None
    if not re.match(r"^DSS\d+$", value): return None
    return value

def get_all_plants_scada(conn):
    plants = []
    with conn.cursor() as cur:
        cur.execute("SELECT DSS_ID FROM DSS_MASTER WHERE DSS_ID IS NOT NULL AND DSS_ID <> '' ORDER BY DSS_ID")
        rows = cur.fetchall()
    for row in rows:
        dss = clean_dss_id(row.get("DSS_ID"))
        if dss: plants.append(dss)
    return sorted(list(set(plants)))

def get_scada_lookup(conn, date):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT DSS_ID, TIMESTAMP, SCADA_POWER_MW FROM DSS_ACTUAL
            WHERE TIMESTAMP >= %s AND TIMESTAMP < DATE_ADD(%s, INTERVAL 1 DAY)
        """, (date, date))
        data_rows = cur.fetchall()
    lookup = {}
    for row in data_rows:
        dss = clean_dss_id(row.get("DSS_ID"))
        if not dss: continue
        ts = row.get("TIMESTAMP")
        if ts is None: continue
        ts = pd.to_datetime(ts, errors="coerce")
        if pd.isna(ts): continue
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        scada_number = pd.to_numeric(row.get("SCADA_POWER_MW"), errors="coerce")
        key = (ts_str, dss)
        lookup[key] = "FULL" if pd.notna(scada_number) else "PARTIAL"
    return lookup, len(data_rows)

def get_monthly_scada_lookup(conn, month):
    start_ts = pd.to_datetime(f"{month}-01")
    end_ts = start_ts + pd.offsets.MonthBegin(1)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT DSS_ID, TIMESTAMP, SCADA_POWER_MW FROM DSS_ACTUAL
            WHERE TIMESTAMP >= %s AND TIMESTAMP < %s
        """, (start_ts.strftime("%Y-%m-%d"), end_ts.strftime("%Y-%m-%d")))
        data_rows = cur.fetchall()
    lookup = {}
    for row in data_rows:
        dss = clean_dss_id(row.get("DSS_ID"))
        if not dss: continue
        ts = pd.to_datetime(row.get("TIMESTAMP"), errors="coerce")
        if pd.isna(ts): continue
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        scada_number = pd.to_numeric(row.get("SCADA_POWER_MW"), errors="coerce")
        lookup[(ts_str, dss)] = "FULL" if pd.notna(scada_number) else "PARTIAL"
    return lookup, len(data_rows), start_ts, end_ts

@app.get("/api/report")
def api_report(date: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        time_slots = pd.date_range(start=f"{date} 00:00:00", end=f"{date} 23:55:00", freq="5min")
        lookup, total_rows = get_scada_lookup(conn, date)
        conn.close()
        
        matrix = {}
        for ts in time_slots:
            ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
            matrix[ts_str] = {plant: lookup.get((ts_str, plant), "NO DATA") for plant in plants}
            
        return {"date": date, "plants": plants, "timestamps": list(matrix.keys()), "total_db_rows": total_rows, "matrix": matrix}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/missing-report")
def missing_report(date: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        time_slots = pd.date_range(start=f"{date} 00:00:00", end=f"{date} 23:55:00", freq="5min")
        total_slots = len(time_slots)
        lookup, total_rows = get_scada_lookup(conn, date)
        conn.close()

        report, total_full, total_partial, total_missing = [], 0, 0, 0
        for plant in plants:
            full, partial, missing, missing_ts = 0, 0, 0, []
            for ts in time_slots:
                ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                status = lookup.get((ts_str, plant), "NO DATA")
                if status == "FULL": full += 1
                elif status == "PARTIAL": partial += 1
                else:
                    missing += 1
                    missing_ts.append(ts_str)
                    
            total_full += full; total_partial += partial; total_missing += missing
            report.append({
                "DSS_ID": plant, "TOTAL_SLOTS": total_slots, "FULL_COUNT": full,
                "PARTIAL_COUNT": partial, "MISSING_COUNT": missing,
                "AVAILABILITY_PERCENT": round((full / total_slots) * 100, 2),
                "MISSING_PERCENT": round((missing / total_slots) * 100, 2),
                "MISSING_TIMESTAMPS": missing_ts
            })
            
        return {"date": date, "total_plants": len(plants), "total_slots": total_slots, "total_db_rows": total_rows, "summary": {"TOTAL_FULL": total_full, "TOTAL_PARTIAL": total_partial, "TOTAL_MISSING": total_missing}, "report": report}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/monthly-missing-report")
def monthly_missing_report(month: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        lookup, total_rows, start_ts, end_ts = get_monthly_scada_lookup(conn, month)
        conn.close()

        time_slots = pd.date_range(start=start_ts.strftime("%Y-%m-%d 00:00:00"), end=(end_ts - pd.Timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S"), freq="5min")
        total_slots = len(time_slots)
        report, total_full, total_partial, total_missing = [], 0, 0, 0

        for plant in plants:
            full, partial, missing, missing_days = 0, 0, 0, {}
            for ts in time_slots:
                ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                day_str = ts.strftime("%Y-%m-%d")
                status = lookup.get((ts_str, plant), "NO DATA")
                if status == "FULL": full += 1
                elif status == "PARTIAL": partial += 1
                else:
                    missing += 1
                    missing_days[day_str] = missing_days.get(day_str, 0) + 1

            total_full += full; total_partial += partial; total_missing += missing
            report.append({
                "DSS_ID": plant, "TOTAL_SLOTS": total_slots, "FULL_COUNT": full,
                "PARTIAL_COUNT": partial, "MISSING_COUNT": missing,
                "AVAILABILITY_PERCENT": round((full / total_slots) * 100, 2),
                "MISSING_PERCENT": round((missing / total_slots) * 100, 2),
                "MISSING_DAYS": missing_days
            })

        return {"month": month, "start_date": start_ts.strftime("%Y-%m-%d"), "end_date": (end_ts - pd.Timedelta(days=1)).strftime("%Y-%m-%d"), "total_plants": len(plants), "total_slots": total_slots, "total_db_rows": total_rows, "summary": {"TOTAL_FULL": total_full, "TOTAL_PARTIAL": total_partial, "TOTAL_MISSING": total_missing}, "report": report}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/missing-report-pdf")
def missing_report_pdf(date: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        time_slots = pd.date_range(start=f"{date} 00:00:00", end=f"{date} 23:55:00", freq="5min")
        total_slots = len(time_slots)
        lookup, total_rows = get_scada_lookup(conn, date)
        conn.close()

        pdf_file = f"missing_scada_report_{date}.pdf"
        pdf_path = os.path.join(os.path.dirname(__file__), pdf_file)
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"SCADA Missing Data Report - {date}", styles["Title"]), Spacer(1, 12)]
        elements.append(Paragraph(f"Date: {date}<br/>Total Plants: {len(plants)}<br/>Total Slots per Plant: {total_slots}<br/>Total DB Rows: {total_rows}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        table_data = [["DSS_ID", "Total Slots", "Full", "Partial", "Missing", "Availability %", "Missing %", "Missing Times"]]
        for plant in plants:
            full, partial, missing, missing_times = 0, 0, 0, []
            for ts in time_slots:
                ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                status = lookup.get((ts_str, plant), "NO DATA")
                if status == "FULL": full += 1
                elif status == "PARTIAL": partial += 1
                else:
                    missing += 1
                    missing_times.append(ts.strftime("%H:%M"))
            
            missing_text = ", ".join(missing_times)
            if len(missing_text) > 180: missing_text = missing_text[:180] + " ..."
            table_data.append([plant, total_slots, full, partial, missing, f"{round((full/total_slots)*100,2)}%", f"{round((missing/total_slots)*100,2)}%", missing_text])

        table = Table(table_data, repeatRows=1, colWidths=[70, 65, 45, 55, 60, 75, 65, 360])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("BACKGROUND", (2, 1), (2, -1), colors.HexColor("#c8e6c9")),
            ("BACKGROUND", (3, 1), (3, -1), colors.HexColor("#fff9c4")), ("BACKGROUND", (4, 1), (4, -1), colors.HexColor("#ffcdd2")),
        ]))
        elements.append(table)
        doc.build(elements)
        return FileResponse(path=pdf_path, filename=pdf_file, media_type="application/pdf")
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
@app.get("/api/missing-report-csv")
def missing_report_csv(date: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        time_slots = pd.date_range(start=f"{date} 00:00:00", end=f"{date} 23:55:00", freq="5min")
        total_slots = len(time_slots)
        lookup, total_rows = get_scada_lookup(conn, date)
        conn.close()

        lines = ["DSS_ID,Total Slots,Full Count,Partial Count,Missing Count,Availability %,Missing %,Missing Timestamps"]
        for plant in plants:
            full, partial, missing, missing_times = 0, 0, 0, []
            for ts in time_slots:
                ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                status = lookup.get((ts_str, plant), "NO DATA")
                if status == "FULL": full += 1
                elif status == "PARTIAL": partial += 1
                else:
                    missing += 1
                    missing_times.append(ts.strftime("%H:%M"))
            
            missing_text = " ".join(missing_times)
            avail_pct = round((full / total_slots) * 100, 2)
            miss_pct = round((missing / total_slots) * 100, 2)
            lines.append(f'{plant},{total_slots},{full},{partial},{missing},{avail_pct}%,{miss_pct}%,"{missing_text}"')

        csv_text = "\n".join(lines)
        fname = f"missing_scada_report_{date}.csv"
        
        return Response(
            content=csv_text,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
@app.get("/api/monthly-missing-report-pdf")
def monthly_missing_report_pdf(month: str = Query(...)):
    try:
        conn = get_scada_connection()
        plants = get_all_plants_scada(conn)
        lookup, total_rows, start_ts, end_ts = get_monthly_scada_lookup(conn, month)
        conn.close()

        time_slots = pd.date_range(start=start_ts.strftime("%Y-%m-%d 00:00:00"), end=(end_ts - pd.Timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S"), freq="5min")
        total_slots = len(time_slots)

        pdf_file = f"monthly_missing_scada_report_{month}.pdf"
        pdf_path = os.path.join(os.path.dirname(__file__), pdf_file)
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Monthly SCADA Missing Data Report - {month}", styles["Title"]), Spacer(1, 12)]
        elements.append(Paragraph(f"Month: {month}<br/>From: {start_ts.strftime('%Y-%m-%d')} To {(end_ts - pd.Timedelta(days=1)).strftime('%Y-%m-%d')}<br/>Total Plants: {len(plants)}<br/>Total Slots per Plant: {total_slots}<br/>Total DB Rows: {total_rows}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        table_data = [["DSS_ID", "Total Slots", "Full", "Partial", "Missing", "Availability %", "Missing %", "Missing Days"]]
        for plant in plants:
            full, partial, missing, missing_days = 0, 0, 0, {}
            for ts in time_slots:
                ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                day_str = ts.strftime("%d-%m")
                status = lookup.get((ts_str, plant), "NO DATA")
                if status == "FULL": full += 1
                elif status == "PARTIAL": partial += 1
                else:
                    missing += 1
                    missing_days[day_str] = missing_days.get(day_str, 0) + 1

            missing_days_text = ", ".join([f"{day}: {count}" for day, count in missing_days.items()])
            if len(missing_days_text) > 220: missing_days_text = missing_days_text[:220] + " ..."
            table_data.append([plant, total_slots, full, partial, missing, f"{round((full/total_slots)*100,2)}%", f"{round((missing/total_slots)*100,2)}%", missing_days_text])

        table = Table(table_data, repeatRows=1, colWidths=[70, 65, 45, 55, 60, 75, 65, 360])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("BACKGROUND", (2, 1), (2, -1), colors.HexColor("#c8e6c9")),
            ("BACKGROUND", (3, 1), (3, -1), colors.HexColor("#fff9c4")), ("BACKGROUND", (4, 1), (4, -1), colors.HexColor("#ffcdd2")),
        ]))
        elements.append(table)
        doc.build(elements)
        return FileResponse(path=pdf_path, filename=pdf_file, media_type="application/pdf")
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    
# =============================================================================
# ── ON-DEMAND FORECASTING INTEGRATION ────────────────────────────────────────
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "forecast_outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

FORECASTER_FILE = BASE_DIR / "combined_forecaster.py"

class ForecastRequest(BaseModel):
    plant: str = "ALL"                 
    date: str                          
    horizon: int = 96                  
    family: Optional[str] = None       
    models_dir: str = "models"
    evaluate: bool = False
    ensemble: bool = False

def _find_col(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    lower = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    return None

def _read_plant_sheet(path: Path, sheet: str) -> Optional[pd.DataFrame]:
    df = pd.read_excel(path, sheet_name=sheet)
    ts_col = _find_col(df, "timestamp", "time", "datetime", "block_time")
    mw_col = _find_col(df, "forecast_mw", "forecast", "predicted_mw", "pred_mw", "mw", "value")
    if ts_col is None or mw_col is None:
        return None
    out = df[[ts_col, mw_col]].copy()
    out.columns = ["timestamp", "forecast_mw"]
    out["timestamp"] = pd.to_datetime(out["timestamp"], errors="coerce")
    out["forecast_mw"] = pd.to_numeric(out["forecast_mw"], errors="coerce")
    out = out.dropna(subset=["timestamp"])
    out["forecast_mw"] = out["forecast_mw"].fillna(0.0)
    return out if not out.empty else None

def _parse_excel_output(path: Path, requested_plant: str = "ALL") -> Dict[str, Any]:
    if not path.exists():
        return {"error": "Forecast finished, but Excel was not created."}

    xls = pd.ExcelFile(path)
    response: Dict[str, Any] = {"sheets": xls.sheet_names, "summary": [], "series": []}

    if "Summary" in xls.sheet_names:
        summary = pd.read_excel(path, sheet_name="Summary")
        response["summary"] = summary.fillna("").to_dict(orient="records")

    plant_sheets = [s for s in xls.sheet_names if s not in ("Summary", "Evaluation")]
    if not plant_sheets:
        response["note"] = "No plant sheets in the Excel — the model produced no forecast rows."
        return response

    frames: Dict[str, pd.DataFrame] = {}
    for s in plant_sheets:
        df = _read_plant_sheet(path, s)
        if df is not None:
            frames[s] = df

    if not frames:
        response["note"] = "Plant sheets found but no recognisable timestamp/MW columns."
        return response

    req = (requested_plant or "ALL").strip()
    if req.upper() == "ALL" and len(frames) > 1:
        combined = pd.concat(frames.values(), ignore_index=True)
        agg = (combined.groupby("timestamp", as_index=False)["forecast_mw"]
                       .sum().sort_values("timestamp"))
        chart_label = f"All Plants (sum of {len(frames)})"
        chart_df = agg
    else:
        if req.upper() != "ALL" and req[:31] in frames:
            chart_label = req[:31]
        else:
            chart_label = max(frames, key=lambda k: frames[k]["forecast_mw"].sum())
        chart_df = frames[chart_label].sort_values("timestamp")

    chart_df = chart_df.copy()
    chart_df["timestamp"] = chart_df["timestamp"].dt.strftime("%Y-%m-%d %H:%M")
    response["series"] = chart_df[["timestamp", "forecast_mw"]].round(3).to_dict(orient="records")
    response["chart_sheet"] = chart_label
    response["n_plants"] = len(frames)
    response["total_points"] = len(chart_df)
    return response

@app.post("/api/run-forecast")
def run_forecast(req: ForecastRequest):
    if not FORECASTER_FILE.exists():
        return JSONResponse({"error": "combined_forecaster.py not found in backend folder."}, status_code=500)

    weather_script = BASE_DIR / "fetch_openmeteo_weather.py"
    if weather_script.exists():
        subprocess.run(["python", str(weather_script), req.date], cwd=str(BASE_DIR))

    output_id = uuid.uuid4().hex[:12]
    output_path = OUTPUT_DIR / f"forecast_{req.date}_{output_id}.xlsx"
    
    cmd = [
        "python", str(FORECASTER_FILE),
        "--date", req.date,
        "--horizon", str(req.horizon),
        "--models-dir", req.models_dir,
        "--output", str(output_path),
    ]

    if req.plant and req.plant.upper() != "ALL":
        cmd.extend(["--plant", req.plant.strip()])
    if req.family:
        cmd.extend(["--family", req.family])
    if req.evaluate:
        cmd.append("--evaluate")
    if req.ensemble:
        cmd.append("--ensemble")

    try:
        result = subprocess.run(cmd, cwd=str(BASE_DIR), text=True, capture_output=True, timeout=1800)
    except subprocess.TimeoutExpired:
        return JSONResponse({"error": "Forecast process timed out."}, status_code=504)

    if result.returncode != 0:
        error_text = (result.stderr or result.stdout or "Forecast failed").strip()
        return JSONResponse({"error": error_text[-4000:]}, status_code=500)

    parsed = _parse_excel_output(output_path, req.plant)
    parsed["download_url"] = f"/api/download-forecast/{output_path.name}"
    parsed["log"] = (result.stdout + "\n" + result.stderr)[-4000:]
    return parsed

@app.get("/api/download-forecast/{filename}")
def download_forecast(filename: str):
    safe_name = Path(filename).name
    path = OUTPUT_DIR / safe_name
    if not path.exists():
        return JSONResponse({"error": "File not found"}, status_code=404)
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=safe_name)

# ── Daily Data Delivery (DSS column-count report over SFTP) ───────────────────
import dss_delivery
from datetime import datetime as _dt

@app.get("/api/daily-delivery")
def get_daily_delivery(date: str = Query(None, description="Run day YYYY-MM-DD; defaults to today")):
    try:
        base_day = _dt.strptime(date, "%Y-%m-%d").date() if date else date.today()
    except ValueError:
        return JSONResponse({"error": "Invalid date. Use YYYY-MM-DD."}, status_code=400)
    return dss_delivery.build_report(base_day, apply_offset=False)

@app.get("/api/daily-delivery/download")
def download_daily_delivery(file: str = Query(..., description="DSS CSV filename")):
    if not dss_delivery.is_valid_dss_filename(file):
        return JSONResponse({"error": "Invalid filename."}, status_code=400)
    try:
        content = dss_delivery.fetch_one(file)
    except Exception as e:
        return JSONResponse({"error": f"Fetch failed: {e}"}, status_code=502)
    if content is None:
        return JSONResponse({"error": "File not found on server."}, status_code=404)
    safe = Path(file).name
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{safe}"'},
    )

# ── Daily Data Delivery: per-company DB load status ───────────────────────────
_DB_EXPECTED_DSS = 212

_DB_REV_MAP = {
    "ISPL":    {"R20": "R16", "R02": "DA2"},
    "QUENEXT": {"R-16": "R16", "Rd2": "DA2"},
}
    
def _rev_label(company, raw):
    raw = str(raw).strip()
    mapped = _DB_REV_MAP.get(company, {}).get(raw)
    if mapped:
        return mapped
    norm = raw.upper().replace("-", "").replace(" ", "")
    for k, v in _DB_REV_MAP.get(company, {}).items():
        if k.upper().replace("-", "").replace(" ", "") == norm:
            return v
    if company == "QUENEXT":
        if raw in ("00_00", "00_25", "00_50", "00_75"):
            return "Day-ahead"
        m = re.match(r"^(\d{2})_00$", raw)
        if m and 1 <= int(m.group(1)) <= 16:
            return f"R{int(m.group(1))} (Intra-day)"
    return raw

_DB_TABLES = {
    "ISPL":    ["dss_forecast_ispl", "dss_forecast_ispl"],
    "QUENEXT": ["dss_forecast_quenext"],
}

def _resolve_table(cursor, candidates):
    for t in candidates:
        try:
            cursor.execute(f"SELECT 1 FROM `{t}` LIMIT 1")
            cursor.fetchall()
            return t
        except Exception:
            continue
    return None

def _db_status_for(company, target_date):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        table = _resolve_table(cursor, _DB_TABLES[company])
        if not table:
            return {"error": f"No DB table found for {company}.", "revisions": []}
        sql = f"""
            SELECT revision_number,
                   COUNT(DISTINCT dss_id)   AS dss_count,
                   COUNT(DISTINCT block_no) AS blocks
            FROM `{table}`
            WHERE forecast_date = %s
            GROUP BY revision_number
            ORDER BY revision_number
        """
        cursor.execute(sql, (target_date,))
        rows = cursor.fetchall()
    except Exception as e:
        return {"error": f"DB query failed for {company}: {e}", "revisions": []}
    finally:
        cursor.close()
        conn.close()

    revisions = []
    for r in rows:
        raw = str(r["revision_number"]).strip()
        cnt = int(r["dss_count"])
        revisions.append({
            "code": raw,
            "label": _rev_label(company, raw),
            "dss": cnt,
            "blocks": int(r["blocks"]),
            "present": cnt > 0,
        })
    return {"error": None, "revisions": revisions}


@app.get("/api/daily-delivery/db")
def daily_delivery_db(date: str = Query(None, description="forecast_date YYYY-MM-DD; default today")):
    try:
        target = _dt.strptime(date, "%Y-%m-%d").date() if date else _dt.now().date()
    except ValueError:
        return JSONResponse({"error": "Invalid date. Use YYYY-MM-DD."}, status_code=400)
    return {
        "date": target.isoformat(),
        "expected": _DB_EXPECTED_DSS,
        "companies": {
            "ISPL":    _db_status_for("ISPL", target),
            "Quenext": _db_status_for("QUENEXT", target),
        },
    }

_DB_REV_RAW = {
    "ISPL":    {"R16": "R20",  "DA2": "R02"},
    "QUENEXT": {"R16": "R-16", "DA2": "Rd2"},
}

def _fmt_csv_time(t):
    if t is None:
        return ""
    if isinstance(t, str):
        return t[:5]
    try:
        secs = int(t.total_seconds())
        return f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}"
    except Exception:
        return str(t)

@app.get("/api/daily-delivery/db-download")
def daily_delivery_db_download(
    company: str = Query(..., description="ISPL or Quenext"),
    revision: str = Query(..., description="R16 or DA2"),
    date: str = Query(..., description="forecast_date YYYY-MM-DD"),
):
    comp = "QUENEXT" if "QUE" in company.upper() else "ISPL"
    raw_norm = revision.upper().replace("-", "").replace(" ", "")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        table = _resolve_table(cursor, _DB_TABLES[comp])
        if not table:
            return JSONResponse({"error": f"No DB table for {comp}."}, status_code=404)
        sql = f"""
            SELECT dss_id, block_no, block_time, timestamp, forecast_mw
            FROM `{table}`
            WHERE forecast_date = %s
              AND REPLACE(REPLACE(UPPER(revision_number), '-', ''), ' ', '') = %s
            ORDER BY block_no, dss_id
        """
        cursor.execute(sql, (date, raw_norm))
        rows = cursor.fetchall()
    except Exception as e:
        return JSONResponse({"error": f"DB query failed: {e}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

    if not rows:
        return JSONResponse(
            {"error": f"No {company} {revision} rows for {date}."}, status_code=404)

    df = pd.DataFrame(rows)
    df["block_time"] = df["block_time"].apply(_fmt_csv_time)
    df["timestamp"] = df["timestamp"].astype(str)

    pivot = df.pivot_table(
        index=["block_no", "block_time", "timestamp"],
        columns="dss_id", values="forecast_mw", aggfunc="first",
    ).reset_index().sort_values("block_no")

    csv_text = pivot.to_csv(index=False)
    fname = f"{comp}_{revision}_{date}.csv"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

from datetime import datetime as _dtime, timedelta as _td

_DAYTIME_GROUPS = [
    ("ISPL",    "R16", "dss_forecast_ispl",    "R20"),
    ("ISPL",    "DA2", "dss_forecast_ispl",    "R02"),
    ("Quenext", "R16", "dss_forecast_quenext", "R16"),
    ("Quenext", "DA2", "dss_forecast_quenext", "RD2"),
]
_DAY_START = "05:30:00"
_DAY_END   = "19:00:00"

def _daytime_slot_labels():
    out, cur, end = _dtime(2000, 1, 1, 5, 30), _dtime(2000, 1, 1, 19, 0)
    while cur <= end:
        out.append(cur.strftime("%H:%M"))
        cur += _td(minutes=15)
    return out

def _fmt_block_time(t):
    if t is None:
        return None
    if isinstance(t, str):
        return t[:5]
    try:
        secs = int(t.total_seconds())
    except AttributeError:
        return str(t)[:5]
    return f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}"

def _run_daytime_group(cursor, company, revision, table, rev_code, slots, date):
    sql = f"""
        SELECT dss_id, block_time, forecast_mw
        FROM {table}
        WHERE forecast_date = %s
          AND REPLACE(REPLACE(UPPER(revision_number), '-', ''), ' ', '') = %s
          AND block_time BETWEEN %s AND %s
        ORDER BY dss_id, block_time
    """
    group = {"company": company, "revision": revision,
             "flagged": [], "flagged_count": 0, "total_dss": 0, "error": None}
    try:
        cursor.execute(sql, (date, rev_code, _DAY_START, _DAY_END))
        rows = cursor.fetchall()
    except Exception as e:
        group["error"] = str(e)
        return group

    per_dss = {}
    for r in rows:
        slot = _fmt_block_time(r["block_time"])
        if slot is None:
            continue
        per_dss.setdefault(r["dss_id"], {})[slot] = r["forecast_mw"]

    group["total_dss"] = len(per_dss)
    for dss in sorted(per_dss.keys()):
        cells = per_dss[dss]
        bad, out_cells = [], {}
        for s in slots:
            v = cells.get(s, None)
            out_cells[s] = None if v is None else round(float(v), 3)
            if (v is None) or (float(v) == 0.0):
                bad.append(s)
        if bad:
            group["flagged"].append({"dss_id": dss, "cells": out_cells, "bad": bad})
    group["flagged_count"] = len(group["flagged"])
    return group

@app.get("/api/daily-delivery/daytime-check")
def daily_delivery_daytime_check(date: str = Query(..., description="Forecast date YYYY-MM-DD")):
    try:
        _dtime.strptime(date, "%Y-%m-%d")
    except ValueError:
        return JSONResponse({"error": "Invalid date. Use YYYY-MM-DD."}, status_code=400)

    slots = _daytime_slot_labels()
    result = {"date": date, "window": {"start": "05:30", "end": "19:00"},
              "slots": slots, "groups": {}}

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        for company, revision, table, rev_code in _DAYTIME_GROUPS:
            result["groups"][f"{company}|{revision}"] = _run_daytime_group(
                cursor, company, revision, table, rev_code, slots, date)
    finally:
        cursor.close()
        conn.close()

    return result

# ── QUENEXT SFTP DEPOSITION REPORT (weekly ON/OFF/MISSING) ───────────────────
@app.get("/api/quenext/deposition-week")
def quenext_deposition_week(week_start: str = Query(..., description="Monday of the week, YYYY-MM-DD")):
    try:
        d = _dtime.strptime(week_start, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse({"error": "Invalid date. Use YYYY-MM-DD."}, status_code=400)

    monday = d - _td(days=d.weekday())
    days = [monday + _td(days=i) for i in range(7)]
    day_strs = [dd.isoformat() for dd in days]

    expected = ["RD1", "RD2", "RD3", "RD4"] + [f"R-{i}" for i in range(1, 17)]

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT forecast_date, revision, status, deposit_time
               FROM quenext_deposition_log
               WHERE forecast_date BETWEEN %s AND %s""",
            (days[0], days[-1]),
        )
        found = {}
        for r in cursor.fetchall():
            dt = r["deposit_time"]
            if dt is not None and not isinstance(dt, str):
                total = int(dt.total_seconds()) if hasattr(dt, "total_seconds") else (dt.hour*3600+dt.minute*60+dt.second)
                dt = f"{total//3600:02d}:{(total%3600)//60:02d}:{total%60:02d}"
            found[(r["forecast_date"].isoformat(), r["revision"])] = {
                "status": r["status"], "time": dt,
            }
    except Exception as e:
        return JSONResponse(
            {"error": f"Read failed: {e}. Has deposition_scan.py run?"},
            status_code=500,
        )
    finally:
        cursor.close()
        conn.close()

    revisions, grid = [], {}
    for rev in expected:
        on = off = miss = 0
        row = {}
        for d_iso in day_strs:
            cell = found.get((d_iso, rev))
            if cell is None:
                row[d_iso] = {"status": "MISSING", "time": None}
                miss += 1
            else:
                row[d_iso] = cell
                if cell["status"] == "ON_TIME":
                    on += 1
                elif cell["status"] == "OFF_TIME":
                    off += 1
                else:
                    miss += 1
        grid[rev] = row
        revisions.append({"revision": rev, "on_time": on, "off_time": off, "missing": miss})

    return {
        "week_start": monday.isoformat(),
        "week_end": days[-1].isoformat(),
        "days": day_strs,
        "revisions": revisions,
        "grid": grid,
    }

# ── SCADA DAILY REPORT (UI page reads scada_report_daily) ────────────────────
@app.get("/api/scada-report")
def scada_report(date: str = Query(..., description="Report date YYYY-MM-DD")):
    try:
        _dtime.strptime(date, "%Y-%m-%d")
    except ValueError:
        return JSONResponse({"error": "Invalid date. Use YYYY-MM-DD."}, status_code=400)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT uss_id, dss_id, pss_name, energy_type,
                      wind_capacity_mw, solar_capacity_mw, total_capacity_mw,
                      visibility_pct, remark, stuck_block_count, stuck_ranges
               FROM scada_report_daily
               WHERE report_date = %s
               ORDER BY (remark = 'OK'), remark, dss_id""",
            (date,),
        )
        rows = cursor.fetchall()
    except Exception as e:
        return JSONResponse(
            {"error": f"Read failed: {e}. Has scada_report_check.py run?"},
            status_code=500,
        )
    finally:
        cursor.close()
        conn.close()

    counts = {
        "total": len(rows),
        "ok": sum(1 for r in rows if r["remark"] == "OK"),
        "no_data": sum(1 for r in rows if "No Data" in (r["remark"] or "")),
        "zero_data": sum(1 for r in rows if r["remark"] == "Zero Data"),
        "stuck_data": sum(1 for r in rows if r["remark"] == "Stuck Data"),
    }
    return {"date": date, "counts": counts, "rows": rows}


# =============================================================================
# ── SCADA DASHBOARD INTEGRATION ──────────────────────────────────────────────
# =============================================================================
import traceback
from datetime import datetime, timedelta

TABLE_NAME = "DSS_ACTUAL"
SLOT_MINUTES = 5
CHECK_SCADA_ONLY = True
FORCE_212_PLANTS = True
DSS_START = 1
DSS_END = 213
SKIP_DSS_IDS = {"DSS00093"}

def dash_sort_key(dss_id):
    try:
        return int(str(dss_id).replace("DSS", ""))
    except Exception:
        return 999999

def dash_build_forced_dss_list():
    plants = []
    for i in range(DSS_START, DSS_END + 1):
        dss = f"DSS{i:05d}"
        if dss not in SKIP_DSS_IDS:
            plants.append(dss)
    return plants

def dash_get_band(availability):
    if availability == 100: return "100%"
    elif 95 <= availability < 100: return "95-99.99%"
    elif 50 <= availability < 95: return "50-95%"
    elif 0 < availability < 50: return "0-50%"
    else: return "0%"

def dash_get_period_range(report_date: str, period: str):
    selected = datetime.strptime(report_date, "%Y-%m-%d")
    if period == "daily":
        return selected, selected + timedelta(days=1)
    elif period == "weekly":
        return selected - timedelta(days=6), selected + timedelta(days=1)
    elif period == "monthly":
        return selected - timedelta(days=29), selected + timedelta(days=1)
    return selected, selected + timedelta(days=1)

def dash_get_days_between(start_dt, end_dt):
    days = []
    current = start_dt
    while current < end_dt:
        days.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    return days

def dash_get_slot_labels():
    base = datetime.strptime("2000-01-01", "%Y-%m-%d")
    slots_per_day = int(24 * 60 / SLOT_MINUTES)
    return [(base + timedelta(minutes=i * SLOT_MINUTES)).strftime("%H:%M") for i in range(slots_per_day)]

def dash_get_all_plants(cur):
    plants_set = set()
    try:
        cur.execute("SELECT DISTINCT DSS_ID FROM DSS_MASTER WHERE DSS_ID IS NOT NULL AND DSS_ID <> ''")
        for r in cur.fetchall(): plants_set.add(str(r["DSS_ID"]))
    except Exception:
        pass

    cur.execute(f"SELECT DISTINCT DSS_ID FROM {TABLE_NAME} WHERE DSS_ID IS NOT NULL AND DSS_ID <> ''")
    for r in cur.fetchall(): plants_set.add(str(r["DSS_ID"]))

    if FORCE_212_PLANTS:
        for dss in dash_build_forced_dss_list(): plants_set.add(dss)

    for dss in SKIP_DSS_IDS: plants_set.discard(dss)
    return sorted(plants_set, key=dash_sort_key)


@app.get("/api/check-db")
def check_db():
    try:
        conn = get_scada_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT DATABASE() AS db_name")
            db = cur.fetchone()
            cur.execute(f"SELECT COUNT(*) AS total_rows FROM {TABLE_NAME}")
            rows = cur.fetchone()
            cur.execute(f"SELECT MIN(TIMESTAMP) AS min_time, MAX(TIMESTAMP) AS max_time FROM {TABLE_NAME}")
            time_range = cur.fetchone()
            try:
                cur.execute("SELECT COUNT(DISTINCT DSS_ID) AS master_plants FROM DSS_MASTER WHERE DSS_ID IS NOT NULL AND DSS_ID <> ''")
                master_count = cur.fetchone()
                master_plants = int(master_count["master_plants"])
            except Exception:
                master_plants = 0
            cur.execute(f"SELECT COUNT(DISTINCT DSS_ID) AS actual_plants FROM {TABLE_NAME} WHERE DSS_ID IS NOT NULL AND DSS_ID <> ''")
            actual_count = cur.fetchone()
        conn.close()

        return {
            "status": "connected", "database": db["db_name"], "table": TABLE_NAME,
            "total_rows": int(rows["total_rows"]), "min_time": str(time_range["min_time"]),
            "max_time": str(time_range["max_time"]), "distinct_plants_in_DSS_MASTER": master_plants,
            "distinct_plants_in_DSS_ACTUAL": int(actual_count["actual_plants"]),
            "forced_dashboard_plants": len(dash_build_forced_dss_list()) if FORCE_212_PLANTS else "disabled",
            "slot_minutes": SLOT_MINUTES
        }
    except Exception as e:
        return JSONResponse({"status": "error", "error": str(e), "traceback": traceback.format_exc()}, status_code=500)


@app.get("/api/plant-list")
def plant_list():
    try:
        conn = get_scada_connection()
        with conn.cursor() as cur:
            plants = dash_get_all_plants(cur)
        conn.close()
        return {"plant_count": len(plants), "plants": plants}
    except Exception as e:
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=500)


@app.get("/api/scada-dashboard")
def dashboard_api(report_date: str = Query(...), period: str = Query("daily")):
    try:
        period = period.lower().strip()
        if period not in ["daily", "weekly", "monthly"]: period = "daily"

        start_dt, end_dt = dash_get_period_range(report_date, period)
        days = dash_get_days_between(start_dt, end_dt)
        start_str, end_str = start_dt.strftime("%Y-%m-%d %H:%M:%S"), end_dt.strftime("%Y-%m-%d %H:%M:%S")

        slot_labels = dash_get_slot_labels()
        slots_per_day, total_days = len(slot_labels), len(days)

        conn = get_scada_connection()
        with conn.cursor() as cur:
            plants = dash_get_all_plants(cur)
            cur.execute(f"SELECT DSS_ID, TIMESTAMP, SCADA_POWER_MW, METER_POWER_MW FROM {TABLE_NAME} WHERE TIMESTAMP >= %s AND TIMESTAMP < %s", (start_str, end_str))
            db_rows = cur.fetchall()
        conn.close()

        total_plants = len(plants)
        expected_slots_per_plant = slots_per_day * total_days
        total_expected_slots = total_plants * expected_slots_per_plant

        data_map = {}
        for row in db_rows:
            dss_id, ts = str(row["DSS_ID"]), row["TIMESTAMP"]
            if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            day, time_label = ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M")
            scada_ok, meter_ok = row["SCADA_POWER_MW"] is not None, row["METER_POWER_MW"] is not None

            if CHECK_SCADA_ONLY: status = "available" if scada_ok else "missing"
            else:
                if scada_ok and meter_ok: status = "available"
                elif scada_ok or meter_ok: status = "partial"
                else: status = "missing"

            data_map.setdefault(dss_id, {}).setdefault(day, {})[time_label] = status

        plant_summary, missing_by_time_total = [], {t: 0 for t in slot_labels}
        daily_available_slots, daily_missing_slots = {day: 0 for day in days}, {day: 0 for day in days}
        daily_missing_plants, daily_full_missing_plants = {day: 0 for day in days}, {day: 0 for day in days}

        for dss_id in plants:
            plant_available, plant_partial, plant_missing = 0, 0, 0
            for day in days:
                day_status = data_map.get(dss_id, {}).get(day, {})
                plant_day_missing = 0
                for t in slot_labels:
                    status = day_status.get(t, "missing")
                    if status == "available":
                        plant_available += 1
                        daily_available_slots[day] += 1
                    elif status == "partial":
                        plant_partial += 1
                        plant_day_missing += 1
                        daily_missing_slots[day] += 1
                        missing_by_time_total[t] += 1
                    else:
                        plant_missing += 1
                        plant_day_missing += 1
                        daily_missing_slots[day] += 1
                        missing_by_time_total[t] += 1

                if plant_day_missing > 0: daily_missing_plants[day] += 1
                if plant_day_missing == slots_per_day: daily_full_missing_plants[day] += 1

            availability = round((plant_available / expected_slots_per_plant) * 100, 2)
            missing_pct = round((plant_missing / expected_slots_per_plant) * 100, 2)
            plant_summary.append({
                "DSS_ID": dss_id, "Total Slots": int(expected_slots_per_plant),
                "Full": int(plant_available), "Partial": int(plant_partial), "Missing": int(plant_missing),
                "Availability %": float(availability), "Missing %": float(missing_pct), "Band": dash_get_band(availability)
            })

        total_available_slots = sum(r["Full"] for r in plant_summary)
        total_missing_slots = sum(r["Missing"] for r in plant_summary)
        avg_availability = round((total_available_slots / total_expected_slots) * 100, 2) if total_expected_slots else 0
        full_day_missing_cases = sum(daily_full_missing_plants.values())
        worst_row = sorted(plant_summary, key=lambda x: (x["Availability %"], -x["Missing"]))[0]

        band_order = ["100%", "95-99.99%", "50-95%", "0-50%", "0%"]
        band_desc = {
            "100%": "All data available", "95-99.99%": "High availability with minor gaps",
            "50-95%": "Partial degradation", "0-50%": "Low availability / major gaps", "0%": "No data available",
        }

        availability_distribution = [{"band": band, "plants": int(sum(1 for r in plant_summary if r["Band"] == band)), "percent": float(round((sum(1 for r in plant_summary if r["Band"] == band) / total_plants) * 100, 2)), "description": band_desc[band]} for band in band_order]
        top_worst = sorted(plant_summary, key=lambda x: (x["Missing %"], x["Missing"]), reverse=True)[:20]
        missing_time_series = [{"time": t, "missing_plants": round(missing_by_time_total[t] / total_days, 2) if total_days else 0} for t in slot_labels]

        slot_distribution = [
            {"name": "0 Missing", "plants": int(sum(1 for r in plant_summary if r["Missing"] == 0))},
            {"name": "1-10 Missing", "plants": int(sum(1 for r in plant_summary if 1 <= r["Missing"] <= 10))},
            {"name": "11-50 Missing", "plants": int(sum(1 for r in plant_summary if 11 <= r["Missing"] <= 50))},
            {"name": "51-287 Missing", "plants": int(sum(1 for r in plant_summary if 51 <= r["Missing"] <= 287))},
            {"name": "288+ Missing", "plants": int(sum(1 for r in plant_summary if r["Missing"] >= 288))},
        ]

        daily_trend = []
        for day in days:
            expected_day_slots = total_plants * slots_per_day
            avail_pct = round((daily_available_slots[day] / expected_day_slots) * 100, 2) if expected_day_slots else 0
            daily_trend.append({"date": day, "availability": float(avail_pct), "missing_slots": int(daily_missing_slots[day]), "missing_plants": int(daily_missing_plants[day]), "full_day_missing_plants": int(daily_full_missing_plants[day])})

        return {
            "report_date": report_date, "period": period, "start_date": start_dt.strftime("%Y-%m-%d"),
            "end_date": (end_dt - timedelta(days=1)).strftime("%Y-%m-%d"), "slot_minutes": int(SLOT_MINUTES),
            "kpis": {
                "total_plants": int(total_plants), "days": int(total_days), "slots_per_day": int(slots_per_day),
                "expected_slots_per_plant": int(expected_slots_per_plant), "total_expected_slots": int(total_expected_slots),
                "total_db_rows": int(len(db_rows)), "total_missing_slots": int(total_missing_slots),
                "full_day_missing_cases": int(full_day_missing_cases), "avg_availability": float(avg_availability),
                "worst_plant": worst_row["DSS_ID"], "worst_plant_availability": float(worst_row["Availability %"]),
                "worst_plant_missing": int(worst_row["Missing"])
            },
            "availability_distribution": availability_distribution, "top_worst": top_worst,
            "missing_time_series": missing_time_series, "slot_distribution": slot_distribution, "daily_trend": daily_trend
        }

    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=500)
# =============================================================================
# ── DATA ENTRY / MAPPING ENDPOINTS ───────────────────────────────────────────
# =============================================================================
from pydantic import BaseModel
from typing import Optional
from datetime import date

class MappingRegistration(BaseModel):
    report_date: date
    sr_no: Optional[int] = None
    voltage_level_kv: Optional[int] = None
    pss_name: str
    district: Optional[str] = None
    zone: Optional[str] = None
    wind_capacity_mw: Optional[float] = 0.0
    solar_capacity_mw: Optional[float] = 0.0
    total_capacity_mw: Optional[float] = 0.0
    uss_id: Optional[str] = None
    dss_id: str
    poi_id: Optional[str] = None
    energy_type: Optional[str] = None
    mapping_type: Optional[str] = None
    billing_group: Optional[str] = None
    other_names: Optional[str] = None
    merged_with_pss: Optional[str] = None

@app.post("/api/mapping/register")
def register_mapping(data: MappingRegistration):
    total_cap = data.total_capacity_mw
    if not total_cap and (data.wind_capacity_mw or data.solar_capacity_mw):
        total_cap = (data.wind_capacity_mw or 0.0) + (data.solar_capacity_mw or 0.0)

    sql = """
        INSERT INTO pss_dss_uss_mapping (
            report_date, sr_no, voltage_level_kv, pss_name, district, zone,
            wind_capacity_mw, solar_capacity_mw, total_capacity_mw, uss_id,
            dss_id, poi_id, energy_type, mapping_type, billing_group,
            other_names, merged_with_pss
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
    """
    
    values = (
        data.report_date, data.sr_no, data.voltage_level_kv, data.pss_name,
        data.district, data.zone, data.wind_capacity_mw, data.solar_capacity_mw,
        total_cap, data.uss_id, data.dss_id, data.poi_id, data.energy_type,
        data.mapping_type, data.billing_group, data.other_names, data.merged_with_pss
    )

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(sql, values)
        conn.commit()
        inserted_id = cursor.lastrowid
        return {"status": "success", "message": "Mapping registered successfully.", "id": inserted_id}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()


# ── VIEW + EDIT pss_dss_uss_mapping ──────────────────────────────────────────
@app.get("/api/mapping/all")
def mapping_all(report_date: str = Query(None, description="Optional YYYY-MM-DD filter")):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if report_date:
            cursor.execute("""
                SELECT id, report_date, sr_no, voltage_level_kv, pss_name, district, zone,
                       wind_capacity_mw, solar_capacity_mw, total_capacity_mw, uss_id, dss_id,
                       poi_id, energy_type, mapping_type, billing_group, other_names, merged_with_pss
                FROM pss_dss_uss_mapping WHERE report_date = %s
                ORDER BY sr_no, pss_name
            """, (report_date,))
        else:
            cursor.execute("""
                SELECT id, report_date, sr_no, voltage_level_kv, pss_name, district, zone,
                       wind_capacity_mw, solar_capacity_mw, total_capacity_mw, uss_id, dss_id,
                       poi_id, energy_type, mapping_type, billing_group, other_names, merged_with_pss
                FROM pss_dss_uss_mapping
                ORDER BY report_date DESC, sr_no, pss_name
            """)
        rows = cursor.fetchall()
        for r in rows:
            if r.get("report_date") is not None:
                r["report_date"] = r["report_date"].isoformat()
        return {"status": "success", "count": len(rows), "rows": rows}
    except Exception as e:
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()


class MappingUpdate(BaseModel):
    report_date: Optional[date] = None
    sr_no: Optional[int] = None
    voltage_level_kv: Optional[int] = None
    pss_name: Optional[str] = None
    district: Optional[str] = None
    zone: Optional[str] = None
    wind_capacity_mw: Optional[float] = None
    solar_capacity_mw: Optional[float] = None
    total_capacity_mw: Optional[float] = None
    uss_id: Optional[str] = None
    dss_id: Optional[str] = None
    poi_id: Optional[str] = None
    energy_type: Optional[str] = None
    mapping_type: Optional[str] = None
    billing_group: Optional[str] = None
    other_names: Optional[str] = None
    merged_with_pss: Optional[str] = None

@app.put("/api/mapping/update/{row_id}")
def mapping_update(row_id: int, data: MappingUpdate):
    fields = data.dict(exclude_unset=True)
    if not fields:
        return JSONResponse({"error": "No fields to update."}, status_code=400)

    if "total_capacity_mw" not in fields and ("wind_capacity_mw" in fields or "solar_capacity_mw" in fields):
        fields["total_capacity_mw"] = (fields.get("wind_capacity_mw") or 0.0) + (fields.get("solar_capacity_mw") or 0.0)

    set_clause = ", ".join(f"{k} = %s" for k in fields.keys())
    values = list(fields.values()) + [row_id]

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"UPDATE pss_dss_uss_mapping SET {set_clause} WHERE id = %s", values)
        conn.commit()
        if cursor.rowcount == 0:
            return JSONResponse({"error": f"No row found with id {row_id}."}, status_code=404)
        return {"status": "success", "message": "Row updated.", "id": row_id}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()


# ── KPTCL: power_market_db.generation_cost_master (separate DB, same server) ──
_KPTCL_COLS = [
    "id", "unit_name", "generator", "station_name", "acronym",
    "contracted_capacity", "valid_from", "valid_to",
    "fixed_cost", "variable_cost", "adjusted_variable_cost",
    "transaction_cost", "trading_margin", "transmission_loss",
    "must_run", "mini_limit", "pool", "plant_type",
    "entire_shared_ent", "entitlement_per_unit", "no_of_units",
    "mintech_per_unit", "mini_limit_percent",
]
_KPTCL_EDITABLE = set(_KPTCL_COLS) - {"id"}  

def _kptcl_conn():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("USE power_market_db")
    cur.close()
    return conn

@app.get("/api/kptcl/all")
def kptcl_all():
    cols = ", ".join(f"`{c}`" for c in _KPTCL_COLS)
    conn = _kptcl_conn()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(f"SELECT {cols} FROM generation_cost_master ORDER BY unit_name, id")
        rows = cursor.fetchall()
        for r in rows:
            for k in ("valid_from", "valid_to"):
                if r.get(k) is not None:
                    r[k] = r[k].isoformat()
            for k, v in list(r.items()):
                if hasattr(v, "is_integer") is False and v.__class__.__name__ == "Decimal":
                    r[k] = float(v)
        return {"status": "success", "count": len(rows), "rows": rows}
    except Exception as e:
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

@app.put("/api/kptcl/update/{row_id}")
def kptcl_update(row_id: int, data: Dict[str, Any]):
    fields = {k: v for k, v in data.items() if k in _KPTCL_EDITABLE}
    if not fields:
        return JSONResponse({"error": "No editable fields to update."}, status_code=400)

    set_clause = ", ".join(f"`{k}` = %s" for k in fields.keys())
    values = list(fields.values()) + [row_id]

    conn = _kptcl_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(f"UPDATE generation_cost_master SET {set_clause} WHERE id = %s", values)
        conn.commit()
        if cursor.rowcount == 0:
            return JSONResponse({"error": f"No row found with id {row_id}."}, status_code=404)
        return {"status": "success", "message": "Row updated.", "id": row_id}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

from datetime import timedelta

# ── GENERIC KPTCL ENDPOINTS FOR NEW TABLES ───────────────────────────────────
ALLOWED_KPTCL_TABLES = {
    "backdown_data": ["id", "report_date", "block_no", "block_time", "si_no", "unit_name", "plant_type", "variable_cost", "minimum_limit", "must_run", "pool", "block_value"],
    "entitlement_data": ["id", "report_date", "block_no", "block_time", "si_no", "unit_name", "plant_type", "variable_cost", "minimum_limit", "must_run", "pool", "block_value"],
    "unit_master": ["id", "unit_name", "generator", "plant_type", "fuel_type", "total_capacity", "contracted_capacity", "min_technical_limit", "time_between_ramp_up_down", "ramp_rate_mw_per_min", "pool", "must_run", "variable_cost", "variable_cost_formula"],
    "market_parameter_master": ["id", "parameter_name", "parameter_value", "parameter_text", "valid_from", "valid_to"],
    "plant_block_data": [
        "id", "timestamp_id", "unit_id", "unit_name", 
        "urs", "sch", "ent", "backdown", 
        "entitlement", "schedule", "calculated_dc", "source_file"
    ]
}

@app.get("/api/kptcl/{table_name}/all")
def kptcl_generic_all(table_name: str):
    if table_name not in ALLOWED_KPTCL_TABLES:
        return JSONResponse({"error": "Invalid table"}, status_code=400)
    
    cols = ", ".join(f"`{c}`" for c in ALLOWED_KPTCL_TABLES[table_name])
    conn = _kptcl_conn()
    cursor = conn.cursor(dictionary=True)
    try:
        limit_clause = " ORDER BY id DESC LIMIT 1500" if "data" in table_name else " ORDER BY id DESC"
        
        cursor.execute(f"SELECT {cols} FROM `{table_name}` {limit_clause}")
        rows = cursor.fetchall()
        
        for r in rows:
            for k, v in list(r.items()):
                if hasattr(v, "isoformat"):
                    r[k] = v.isoformat()
                elif isinstance(v, timedelta):
                    secs = int(v.total_seconds())
                    r[k] = f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}"
                elif hasattr(v, "is_integer") is False and v.__class__.__name__ == "Decimal":
                    r[k] = float(v)
                    
        return {"status": "success", "count": len(rows), "rows": rows}
    except Exception as e:
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

@app.put("/api/kptcl/{table_name}/update/{row_id}")
def kptcl_generic_update(table_name: str, row_id: int, data: Dict[str, Any]):
    if table_name not in ALLOWED_KPTCL_TABLES:
        return JSONResponse({"error": "Invalid table"}, status_code=400)

    editable_cols = set(ALLOWED_KPTCL_TABLES[table_name]) - {"id"}
    fields = {k: v for k, v in data.items() if k in editable_cols}
    
    if not fields:
        return JSONResponse({"error": "No editable fields to update."}, status_code=400)

    set_clause = ", ".join(f"`{k}` = %s" for k in fields.keys())
    values = list(fields.values()) + [row_id]

    conn = _kptcl_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(f"UPDATE `{table_name}` SET {set_clause} WHERE id = %s", values)
        conn.commit()
        if cursor.rowcount == 0:
            return JSONResponse({"error": f"No row found with id {row_id}."}, status_code=404)
        return {"status": "success", "message": "Row updated.", "id": row_id}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

_KPTCL_HASH_TABLES = {
    "generation_cost_master", "backdown_data", "ent_data",
    "entitlement_data", "sch_data", "unit_master", "market_parameter_master",
}

def _compute_row_hash(fields: dict) -> str:
    parts = [f"{k}={'' if fields[k] is None else fields[k]}" for k in sorted(fields.keys())]
    joined = "|".join(parts)
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()

def _table_has_column(cursor, table_name, column):
    cursor.execute(
        """SELECT 1 FROM information_schema.columns
           WHERE table_schema = 'power_market_db' AND table_name = %s AND column_name = %s
           LIMIT 1""",
        (table_name, column),
    )
    return cursor.fetchone() is not None

@app.post("/api/kptcl/{table_name}/insert")
def kptcl_generic_insert(table_name: str, data: Dict[str, Any]):
    if table_name not in ALLOWED_KPTCL_TABLES:
        return JSONResponse({"error": "Invalid table"}, status_code=400)

    insertable_cols = [c for c in ALLOWED_KPTCL_TABLES[table_name] if c != "id"]
    fields = {k: v for k, v in data.items() if k in insertable_cols and v not in (None, "")}

    if not fields:
        return JSONResponse({"error": "No values provided to insert."}, status_code=400)

    conn = _kptcl_conn()
    cursor = conn.cursor()
    try:
        insert_fields = dict(fields)
        if _table_has_column(cursor, table_name, "row_hash"):
            insert_fields["row_hash"] = _compute_row_hash(fields)

        col_clause = ", ".join(f"`{k}`" for k in insert_fields.keys())
        placeholders = ", ".join(["%s"] * len(insert_fields))
        values = list(insert_fields.values())

        cursor.execute(
            f"INSERT INTO `{table_name}` ({col_clause}) VALUES ({placeholders})",
            values,
        )
        conn.commit()
        return {"status": "success", "message": "Row added.", "id": cursor.lastrowid}
    except Exception as e:
        conn.rollback()
        msg = str(e)
        if "Duplicate entry" in msg and "row_hash" in msg:
            return JSONResponse({"error": "An identical row already exists."}, status_code=409)
        return JSONResponse({"error": f"Database error: {msg}"}, status_code=500)
    finally:
        cursor.close()
        conn.close()

# ── KPTCL EXCEL UPLOAD ────────────────────────────────────────────────────────
@app.post("/api/kptcl/upload-excel")
async def kptcl_upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse({"error": "Only .xlsx files are accepted."}, status_code=400)
    content = await file.read()
    try:
        from upload_kptcl import process_upload
        from database import DB_CONFIG
        summary = process_upload(content, file.filename, DB_CONFIG)
        return summary
    except Exception as e:
        return JSONResponse({"error": f"Upload processing failed: {type(e).__name__}: {e}"}, status_code=500)

# ── PLANT BLOCK DATA: view by date (joined with plant_block_key and calculated_values)
@app.get("/api/kptcl/block-data")
def kptcl_block_data(date: str = Query(None, description="Single date YYYY-MM-DD (backward compat)"),
                     date_from: str = Query(None, description="Range start YYYY-MM-DD"),
                     date_to: str = Query(None, description="Range end YYYY-MM-DD"),
                     limit: int = Query(50000)):
    """Return plant_block_data joined with plant_block_key and calculated_values."""
    conn = _kptcl_conn()
    cursor = conn.cursor(dictionary=True)
    try:
        if date_from and date_to:
            cursor.execute("""
                SELECT p.id, k.report_date, k.block_no, k.block_time,
                       p.unit_name, p.urs, p.sch, p.ent, p.backdown,
                       p.entitlement, p.schedule, p.calculated_dc,
                       cv.must_run_value, cv.minimum_capacity,
                       cv.capacity_over_min, cv.capacity_over_min_act_sch
                FROM plant_block_data p
                JOIN plant_block_key k ON k.block_id = p.timestamp_id
                LEFT JOIN calculated_values cv ON cv.timestamp_id = p.timestamp_id AND cv.unit_id = p.unit_id
                WHERE k.report_date BETWEEN %s AND %s
                ORDER BY k.report_date, k.block_no, p.unit_name
                LIMIT %s
            """, (date_from, date_to, limit))
        else:
            d = date or date_from or date_to
            if not d:
                return JSONResponse({"error": "Provide date or date_from+date_to."}, status_code=400)
            cursor.execute("""
                SELECT p.id, k.report_date, k.block_no, k.block_time,
                       p.unit_name, p.urs, p.sch, p.ent, p.backdown,
                       p.entitlement, p.schedule, p.calculated_dc,
                       cv.must_run_value, cv.minimum_capacity,
                       cv.capacity_over_min, cv.capacity_over_min_act_sch
                FROM plant_block_data p
                JOIN plant_block_key k ON k.block_id = p.timestamp_id
                LEFT JOIN calculated_values cv ON cv.timestamp_id = p.timestamp_id AND cv.unit_id = p.unit_id
                WHERE k.report_date = %s
                ORDER BY k.block_no, p.unit_name
                LIMIT %s
            """, (d, limit))
        rows = cursor.fetchall()
        for r in rows:
            if r.get("report_date"):
                r["report_date"] = r["report_date"].isoformat()
            if r.get("block_time"):
                bt = r["block_time"]
                total = int(bt.total_seconds()) if hasattr(bt, "total_seconds") else 0
                r["block_time"] = f"{total//3600:02d}:{(total%3600)//60:02d}"
            
            # Format the numeric values, including the newly joined ones
            for k in ("urs","sch","ent","backdown","entitlement","schedule","calculated_dc",
                      "must_run_value", "minimum_capacity", "capacity_over_min", "capacity_over_min_act_sch"):
                if r.get(k) is not None and not isinstance(r[k], (int, float)):
                    r[k] = float(r[k])

        # fetch unit metadata
        cursor.execute("""
            SELECT unit_name, plant_type, variable_cost, must_run, min_technical_limit, pool
            FROM unit_master
        """)
        unit_meta = {}
        for u in cursor.fetchall():
            unit_meta[u["unit_name"]] = {
                "plant_type": u.get("plant_type") or "",
                "variable_cost": float(u["variable_cost"]) if u.get("variable_cost") is not None else 0,
                "must_run": "TRUE" if u.get("must_run") else "FALSE",
                "minimum_limit": float(u["min_technical_limit"]) if u.get("min_technical_limit") is not None else 0,
                "pool": u.get("pool") or "",
            }

        return {"status": "success", "count": len(rows), "rows": rows, "unit_meta": unit_meta}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        cursor.close()
        conn.close()

@app.get("/api/kptcl/block-data/dates")
def kptcl_block_data_dates():
    conn = _kptcl_conn()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT DISTINCT report_date FROM plant_block_key ORDER BY report_date DESC LIMIT 100")
        dates = [r[0].isoformat() for r in cursor.fetchall()]
        return {"dates": dates}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        cursor.close()
        conn.close()

@app.get("/api/kptcl/block-data/download")
def kptcl_block_data_download(date: str = Query(None),
                               date_from: str = Query(None),
                               date_to: str = Query(None)):
    """Download plant_block_data and calculated_values as a multi-sheet Excel."""
    conn = _kptcl_conn()
    cursor = conn.cursor(dictionary=True)
    try:
        if date_from and date_to:
            cursor.execute("""
                SELECT k.report_date, k.block_no, k.block_time,
                       p.unit_name, p.urs, p.sch, p.ent, p.backdown,
                       p.entitlement, p.schedule, p.calculated_dc,
                       cv.must_run_value, cv.minimum_capacity,
                       cv.capacity_over_min, cv.capacity_over_min_act_sch
                FROM plant_block_data p
                JOIN plant_block_key k ON k.block_id = p.timestamp_id
                LEFT JOIN calculated_values cv ON cv.timestamp_id = p.timestamp_id AND cv.unit_id = p.unit_id
                WHERE k.report_date BETWEEN %s AND %s
                ORDER BY k.report_date, k.block_no, p.unit_name
            """, (date_from, date_to))
            fname = f"block_data_{date_from}_to_{date_to}.xlsx"
        else:
            d = date or date_from or date_to
            if not d:
                return JSONResponse({"error": "Provide date or date_from+date_to."}, status_code=400)
            cursor.execute("""
                SELECT k.report_date, k.block_no, k.block_time,
                       p.unit_name, p.urs, p.sch, p.ent, p.backdown,
                       p.entitlement, p.schedule, p.calculated_dc,
                       cv.must_run_value, cv.minimum_capacity,
                       cv.capacity_over_min, cv.capacity_over_min_act_sch
                FROM plant_block_data p
                JOIN plant_block_key k ON k.block_id = p.timestamp_id
                LEFT JOIN calculated_values cv ON cv.timestamp_id = p.timestamp_id AND cv.unit_id = p.unit_id
                WHERE k.report_date = %s
                ORDER BY k.block_no, p.unit_name
            """, (d,))
            fname = f"block_data_{d}.xlsx"
        rows = cursor.fetchall()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        cursor.close()
        conn.close()

    if not rows:
        return JSONResponse({"error": "No data for selected dates."}, status_code=404)

    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    units = sorted(set(r["unit_name"] for r in rows))
    pivot = {}
    for r in rows:
        key = (str(r["report_date"]), r["block_no"], str(r.get("block_time", "")))
        if key not in pivot:
            pivot[key] = {}
        pivot[key][r["unit_name"]] = r
    sorted_keys = sorted(pivot.keys(), key=lambda k: (k[0], int(k[1])))

    # Added the 4 calculated fields as new sheets
    sheets = [
        ("URS", "urs"), ("SCH", "sch"), ("ENT", "ent"),
        ("Backdown", "backdown"), ("Entitlement", "entitlement"),
        ("Schedule", "schedule"), ("Calculated-DC", "calculated_dc"),
        ("MustRun", "must_run_value"), ("Minimum_Capacity", "minimum_capacity"),
        ("CapacityOverMin", "capacity_over_min"), ("CapOverMinActSch", "capacity_over_min_act_sch")
    ]

    conn2 = _kptcl_conn()
    cur2 = conn2.cursor(dictionary=True)
    cur2.execute("SELECT unit_name, plant_type, variable_cost, must_run, min_technical_limit, pool FROM unit_master")
    umeta = {}
    for u in cur2.fetchall():
        umeta[u["unit_name"]] = {
            "Plant_Type": u.get("plant_type") or "",
            "Variable_cost": float(u["variable_cost"]) if u.get("variable_cost") is not None else 0,
            "Must Run": "TRUE" if u.get("must_run") else "FALSE",
            "Minimum_limit": float(u["min_technical_limit"]) if u.get("min_technical_limit") is not None else 0.0,
            "pool": u.get("pool") or "",
        }
    cur2.close()
    conn2.close()

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="1a2535", end_color="1a2535", fill_type="solid")
    meta_font = Font(italic=True, color="888888", size=9)

    meta_rows = ["Plant_Type", "Variable_cost", "Must Run", "Minimum_limit", "pool"]

    for sheet_name, field in sheets:
        ws = wb.create_sheet(title=sheet_name)
        for ri, mk in enumerate(meta_rows, 1):
            ws.cell(row=ri, column=1, value="").font = meta_font
            ws.cell(row=ri, column=2, value=mk).font = meta_font
            for ui, unit in enumerate(units, 3):
                val = umeta.get(unit, {}).get(mk, "")
                ws.cell(row=ri, column=ui, value=val).font = meta_font

        headers = ["Date", "Block_no"] + units
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, key in enumerate(sorted_keys, 7):
            dt, bno, btime = key
            ws.cell(row=ri, column=1, value=dt)
            ws.cell(row=ri, column=2, value=int(bno))
            for ui, unit in enumerate(units, 3):
                r = pivot[key].get(unit)
                val = r.get(field) if r else None
                if val is not None:
                    ws.cell(row=ri, column=ui, value=float(val))
                else:
                    ws.cell(row=ri, column=ui, value=0)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── SOLAR / WIND STATIC DETAILS ──────────────────────────────────────────────
class SolarStaticDetail(BaseModel):
    dss_id: str
    uss_id: Optional[str] = None
    uss_name: Optional[str] = None
    dss_name: Optional[str] = None
    capacity_gen: Optional[float] = None
    capacity_inj: Optional[float] = None
    energy_type: Optional[str] = "SOLAR"
    dss_status: Optional[str] = None
    location_id: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None

@app.post("/api/solar-static/register")
def register_solar_static(data: SolarStaticDetail):
    sql = """
        INSERT INTO solar_static_details (
            dss_id, uss_id, uss_name, dss_name, capacity_gen, capacity_inj,
            energy_type, dss_status, location_id, latitude, longitude
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    values = (
        data.dss_id, data.uss_id, data.uss_name, data.dss_name,
        data.capacity_gen, data.capacity_inj, data.energy_type,
        data.dss_status, data.location_id, data.latitude, data.longitude
    )
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(sql, values)
        conn.commit()
        return {"status": "success", "message": "Solar detail saved.", "id": data.dss_id}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close(); conn.close()


class WindStaticDetail(BaseModel):
    dss_id: str
    uss_id: Optional[str] = None
    uss_name: Optional[str] = None
    dss_name: Optional[str] = None
    wtg_id: Optional[str] = None
    wtg_name: Optional[str] = None
    capacity_gen: Optional[float] = None
    capacity_inj: Optional[float] = None
    energy_type: Optional[str] = "WIND"
    dss_status: Optional[str] = None
    location_id: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    turbine_id: Optional[str] = None
    manufacturer_name: Optional[str] = None
    model: Optional[str] = None
    capacity: Optional[float] = None

@app.post("/api/wind-static/register")
def register_wind_static(data: WindStaticDetail):
    sql = """
        INSERT INTO us_ds_static_details (
            dss_id, uss_id, uss_name, dss_name, wtg_id, wtg_name,
            capacity_gen, capacity_inj, energy_type, dss_status, location_id,
            latitude, longitude, turbine_id, manufacturer_name, model, capacity
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    values = (
        data.dss_id, data.uss_id, data.uss_name, data.dss_name, data.wtg_id,
        data.wtg_name, data.capacity_gen, data.capacity_inj, data.energy_type,
        data.dss_status, data.location_id, data.latitude, data.longitude,
        data.turbine_id, data.manufacturer_name, data.model, data.capacity
    )
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(sql, values)
        conn.commit()
        return {"status": "success", "message": "Wind detail saved.", "id": cursor.lastrowid}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": f"Database error: {str(e)}"}, status_code=500)
    finally:
        cursor.close(); conn.close()

from dashboard_api import router as dashboard_router
app.include_router(dashboard_router)

from settlement_api import router as settlement_router
app.include_router(settlement_router)

ui_path = os.path.join(os.path.dirname(__file__), "ui")
if os.path.exists(ui_path):
    app.mount("/", StaticFiles(directory=ui_path, html=True), name="ui")
else:
    print(f"WARNING: UI directory not found at {ui_path}. Static files will not be served.")